#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
启贤托育AI税务助手 - Web 界面（支持真实 DeepSeek AI）
运行：streamlit run tax_web_app.py

API Key 配置优先级：
  1. 侧边栏手动输入（最高）
  2. .streamlit/secrets.toml 或 Streamlit Cloud secrets
  3. .env 文件中的 DEEPSEEK_API_KEY
  4. 未配置则使用模拟模式
"""

import sys
import os
import json
import requests
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import streamlit as st
import pandas as pd
from datetime import datetime
from dotenv import load_dotenv

# 启动时加载 .env 文件
load_dotenv()

# 导入计算参数
from tax_calculator import (
    SOCIAL_INSURANCE_ACTUAL,
    SOCIAL_INSURANCE_COMPANY,
    BASIC_DEDUCTION,
    calc_corporate_income_tax_quarterly,
    format_corporate_tax_report,
    classify_bank_transaction,
    generate_profit_statement,
    validate_quarterly_declaration,
    calc_vat_and_surcharge,
    get_tax_policy_summary,
    calc_disabled_employment_fund,
    calc_stamp_duty,
    validate_salary_data,
)

# ===============================================
#  PDF 生成工具
# ===============================================

try:
    from fpdf import FPDF
    _HAS_FPDF = True
except ImportError:
    _HAS_FPDF = False


def make_pdf(title: str, body_lines: list, filename: str) -> bytes | None:
    """生成 PDF 文件字节流，fpdf2 不可用时返回 None"""
    if not _HAS_FPDF:
        return None
    try:
        pdf = FPDF()
        pdf.add_page()
        # 使用内置字体（无需额外字体文件）
        pdf.set_auto_page_break(auto=True, margin=15)
        # 标题
        pdf.set_font("Helvetica", "B", 14)
        pdf.cell(0, 10, title, ln=True, align="C")
        pdf.ln(6)
        # 正文
        pdf.set_font("Helvetica", "", 9)
        for line in body_lines:
            # 处理中文：用 ASCII 替代方案
            safe_line = line.encode("ascii", errors="replace").decode("ascii")
            pdf.multi_cell(0, 5, safe_line)
        return bytes(pdf.output())
    except Exception:
        return None


def make_pdf_with_dataframe(title: str, df, summary_lines: list, filename: str) -> bytes | None:
    """生成含表格的 PDF"""
    if not _HAS_FPDF:
        return None
    try:
        pdf = FPDF()
        pdf.add_page("L")  # 横向
        pdf.set_auto_page_break(auto=True, margin=10)
        pdf.set_font("Helvetica", "B", 12)
        pdf.cell(0, 8, title, ln=True, align="C")
        pdf.ln(4)

        # 表头
        cols = list(df.columns)
        col_w = (pdf.w - 20) / len(cols)
        pdf.set_font("Helvetica", "B", 7)
        for c in cols:
            safe = str(c).encode("ascii", errors="replace").decode("ascii")
            pdf.cell(col_w, 6, safe[:12], border=1, align="C")
        pdf.ln()

        # 数据行
        pdf.set_font("Helvetica", "", 7)
        for _, row in df.head(30).iterrows():
            for c in cols:
                val = str(row[c]).encode("ascii", errors="replace").decode("ascii")
                pdf.cell(col_w, 5, val[:15], border=1, align="C")
            pdf.ln()

        pdf.ln(6)
        pdf.set_font("Helvetica", "", 9)
        for line in summary_lines:
            safe = line.encode("ascii", errors="replace").decode("ascii")
            pdf.multi_cell(0, 5, safe)

        return bytes(pdf.output())
    except Exception:
        return None

# ===============================================
#  DeepSeek AI 配置
# ===============================================

DEEPSEEK_API_URL = "https://api.deepseek.com/chat/completions"
DEEPSEEK_MODEL = "deepseek-chat"

# 从 .env 读取默认 Key
ENV_API_KEY = os.environ.get("DEEPSEEK_API_KEY", "")


def get_api_key():
    """获取当前生效的 API Key（手动输入 > st.secrets > .env）"""
    # 1. 侧边栏手动输入（最高优先级）
    manual_key = st.session_state.get("deepseek_api_key_manual", "")
    if manual_key:
        return manual_key, "manual"
    # 2. Streamlit Cloud secrets
    try:
        if "DEEPSEEK_API_KEY" in st.secrets:
            return st.secrets["DEEPSEEK_API_KEY"], "secrets"
    except Exception:
        pass
    # 3. .env 文件
    if ENV_API_KEY:
        return ENV_API_KEY, "env"
    return "", "none"


def ask_deepseek(prompt: str, system_prompt: str = None) -> str:
    """调用 DeepSeek API"""
    api_key, _ = get_api_key()
    if not api_key:
        return "[未配置 API Key，跳过 AI 生成]"

    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json",
    }
    messages = []
    if system_prompt:
        messages.append({"role": "system", "content": system_prompt})
    messages.append({"role": "user", "content": prompt})

    payload = {
        "model": DEEPSEEK_MODEL,
        "messages": messages,
        "temperature": 0.3,
        "max_tokens": 1500,
    }

    try:
        resp = requests.post(
            DEEPSEEK_API_URL, headers=headers, json=payload, timeout=30
        )
        resp.raise_for_status()
        return resp.json()["choices"][0]["message"]["content"]
    except Exception as e:
        return f"[AI 调用失败: {e}]"


# ===============================================
#  核心计算函数
# ===============================================

def calc_one_employee(
    name, gross_salary, si_base, si_personal,
    special_total, child_edu, infant_care, elderly_care,
    housing_fund_personal=0.0,
) -> dict:
    """计算单名员工"""
    taxable_income = (
        gross_salary
        - si_personal
        - housing_fund_personal
        - BASIC_DEDUCTION
        - special_total
    )

    if taxable_income <= 0:
        tax = 0.0
        taxable_income = 0.0
    else:
        if taxable_income <= 3000:
            rate, deduction = 0.03, 0
        elif taxable_income <= 12000:
            rate, deduction = 0.10, 210
        elif taxable_income <= 25000:
            rate, deduction = 0.20, 1410
        elif taxable_income <= 35000:
            rate, deduction = 0.25, 2660
        elif taxable_income <= 55000:
            rate, deduction = 0.30, 4410
        elif taxable_income <= 80000:
            rate, deduction = 0.35, 7160
        else:
            rate, deduction = 0.45, 15160

        tax = taxable_income * rate - deduction

    net_salary = gross_salary - si_personal - housing_fund_personal - round(tax, 2)

    si_company = (
        si_base * SOCIAL_INSURANCE_COMPANY["pension"]
        + si_base * SOCIAL_INSURANCE_COMPANY["medical"]
        + si_base * SOCIAL_INSURANCE_COMPANY["unemployment"]
        + si_base * SOCIAL_INSURANCE_COMPANY["injury"]
    )
    total_cost = gross_salary + si_company

    return {
        "姓名": name,
        "税前工资": gross_salary,
        "个人社保": si_personal,
        "专项附加扣除": special_total,
        "子女教育": child_edu,
        "婴幼儿照护": infant_care,
        "赡养老人": elderly_care,
        "应税收入": round(taxable_income, 2),
        "应纳税额": round(tax, 2),
        "实发工资": round(net_salary, 2),
        "公司社保承担": round(si_company, 2),
        "公司用人总成本": round(total_cost, 2),
    }


# ===============================================
#  AI 申报说明生成
# ===============================================

def generate_tax_report_ai(results: list) -> str:
    """生成个税申报说明（真实 AI 或模拟）"""
    now = datetime.now()
    api_key, _ = get_api_key()
    use_ai = bool(api_key)

    if use_ai:
        rows_text = ""
        for r in results:
            rows_text += (
                f"员工 {r['姓名']}：税前工资 {r['税前工资']} 元，"
                f"个人社保 {r['个人社保']} 元，"
                f"专项附加扣除 {r['专项附加扣除']} 元"
                f"（子女教育 {r['子女教育']} 元，婴幼儿照护 {r['婴幼儿照护']} 元，"
                f"赡养老人 {r['赡养老人']} 元），"
                f"应税收入 {r['应税收入']} 元，应纳税额 {r['应纳税额']} 元，"
                f"实发工资 {r['实发工资']} 元。\n"
            )

        company_si_total = sum(r["公司社保承担"] for r in results)
        total_tax = sum(r["应纳税额"] for r in results)
        total_cost = sum(r["公司用人总成本"] for r in results)

        prompt = f"""你是一位专业的税务顾问，请为以下企业{now.year}年{now.month}月的个税及社保申报撰写一份专业的申报说明。

## 员工数据
{rows_text}
## 汇总数据
- 公司承担社保总额：{company_si_total} 元
- 全体员工应纳税额合计：{total_tax} 元
- 公司用人总成本：{total_cost} 元

## 要求
1. 以"湖北启贤托儿所有限公司 {now.year}年{now.month}月 税务申报说明"为标题
2. 分四个部分：一、申报概况；二、员工个税明细；三、社保缴纳说明；四、申报注意事项
3. 语气专业、简洁，适合财务提交给税务局或留存备案
4. 提醒用户核对专项附加扣除信息是否已及时更新（个税APP）
5. 说明社保基数如有调整请以社保局核定为准
6. 总字数控制在 500-800 字
7. 用中文输出，不要输出英文
"""

        ai_result = ask_deepseek(
            prompt,
            system_prompt="你是一位专业的税务顾问，擅长撰写企业税务申报说明。"
        )
        if not ai_result.startswith("["):
            return ai_result

    # 模拟模式
    has_tax = any(r["应纳税额"] > 0 for r in results)
    lines = [f"【{now.year}年{now.month}月个税申报说明】", ""]
    lines.append(f"本月公司共有 {len(results)} 名员工需进行个税申报。")
    if has_tax:
        total_tax = sum(r["应纳税额"] for r in results)
        lines.append(f"本月应纳个税合计 {total_tax:.2f} 元，请及时在自然人电子税务局（扣缴端）完成申报缴税。")
    else:
        lines.append("经计算，本月所有员工应税收入均为 0 元，无需缴纳个税。请在自然人电子税务局进行零申报操作。")
    lines.append("")
    lines.append("【扣除项说明】")
    for r in results:
        lines.append(
            f"  {r['姓名']}：社保扣除 {r['个人社保']} 元，"
            f"专项附加扣除 {r['专项附加扣除']} 元"
            f"（子女教育{r['子女教育']}+婴幼儿{r['婴幼儿照护']}+赡养老人{r['赡养老人']}）。"
        )
    lines.append("")
    lines.append("【注意事项】")
    lines.append("  1. 请核实员工专项附加扣除信息是否最新；")
    lines.append("  2. 社保基数如有调整，请及时更新系统参数；")
    lines.append("  3. 零申报也需按时提交，避免产生逾期记录。")
    lines.append("")
    lines.append("—— 由 启贤托育AI税务助手 自动生成")
    return "\n".join(lines)


def generate_social_report_ai(results: list) -> str:
    """生成社保申报说明（真实 AI 或模拟）"""
    now = datetime.now()
    api_key, _ = get_api_key()
    use_ai = bool(api_key)

    if use_ai:
        rows_text = ""
        for r in results:
            rows_text += (
                f"  员工{r['姓名']}：缴费基数 5000 元，"
                f"公司社保承担 {r['公司社保承担']} 元，"
                f"个人社保 {r['个人社保']} 元。\n"
            )
        total_si = sum(r["公司社保承担"] for r in results)

        prompt = f"""请为湖北启贤托儿所有限公司生成 {now.year}年{now.month}月 的社保申报操作说明。

社保数据：
{rows_text}
汇总：公司承担社保合计 {total_si} 元。

要求：
1. 说明社保缴纳明细和公司承担部分
2. 提供操作指引（登录湖北政务服务网，进入单位社保申报模块）
3. 提醒申报截止时间和注意事项
4. 语气专业，200-300 字，用中文输出
"""

        ai_result = ask_deepseek(prompt)
        if not ai_result.startswith("["):
            return ai_result

    # 模拟模式
    total_si = sum(r["公司社保承担"] for r in results)
    lines = [
        f"【{now.year}年{now.month}月社保申报说明】",
        "",
        f"本月需为 {len(results)} 名员工缴纳社保，公司承担部分合计 {total_si:.2f} 元。",
        "",
        "【缴费明细】",
    ]
    for r in results:
        lines.append(
            f"  {r['姓名']}：缴费基数 5000 元，"
            f"公司承担 {r['公司社保承担']} 元，"
            f"个人承担 {r['个人社保']} 元。"
        )
    lines.append("")
    lines.append("【操作指引】")
    lines.append("  1. 登录「湖北政务服务网」或「武汉社保申报系统」；")
    lines.append("  2. 进入「单位社保申报」模块，核对人员名单；")
    lines.append("  3. 确认缴费基数无误后提交申报；")
    lines.append("  4. 缴费成功后留存缴费凭证备查。")
    lines.append("")
    lines.append("【注意事项】")
    lines.append("  社保申报截止时间为每月 25 日，请提前办理。")
    lines.append("")
    lines.append("—— 由 启贤托育AI税务助手 自动生成")
    return "\n".join(lines)


# ===============================================
#  生成上传模板（内存中）
# ===============================================

def get_template_df():
    """返回示范用的上传模板 DataFrame"""
    return pd.DataFrame([
        {
            "姓名": "员工A",
            "税前工资": 10522,
            "社保基数": 5000,
            "个人社保实缴": 522,
            "专项附加扣除": 5000,
            "子女教育": 2000,
            "婴幼儿照护": 2000,
            "赡养老人": 1000,
        },
        {
            "姓名": "员工B",
            "税前工资": 8000,
            "社保基数": 5000,
            "个人社保实缴": 522,
            "专项附加扣除": 0,
            "子女教育": 0,
            "婴幼儿照护": 0,
            "赡养老人": 0,
        },
    ])


# ===============================================
#  年报导入模板生成
# ===============================================

ANNUAL_TEMPLATE_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "年报数据导入模板.xlsx")


def gen_annual_report_template_bytes() -> bytes:
    """生成「年报数据导入」Excel 模板（3个Sheet），返回字节流"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    # ── 通用样式 ──
    header_font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    body_font = Font(name="微软雅黑", size=10)
    body_align = Alignment(horizontal="left", vertical="center")
    num_align = Alignment(horizontal="right", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    section_font = Font(name="微软雅黑", bold=True, size=11, color="1E40AF")
    hint_font = Font(name="微软雅黑", size=9, color="6B7280")

    def style_header(ws, row, cols):
        for c in range(1, cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

    def style_body_row(ws, row, cols, is_num=False):
        for c in range(1, cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = body_font
            cell.alignment = num_align if (is_num and c >= 2) else body_align
            cell.border = thin_border

    # ============================================================
    #  Sheet 1：年报汇总
    # ============================================================
    ws1 = wb.active
    ws1.title = "年报汇总"

    ws1.merge_cells("A1:D1")
    ws1.cell(row=1, column=1, value="湖北启贤托儿所有限公司 · 2025年度年报数据导入模板").font = Font(name="微软雅黑", bold=True, size=14, color="1E3A8A")
    ws1.cell(row=2, column=1, value="填写说明：仅需在「数值」列填入实际金额；灰色行无需填写；填完后切换到「员工信息」Sheet 填写员工明细").font = hint_font
    ws1.merge_cells("A2:D2")

    headers1 = ["序号", "项目名称", "数值", "单位 / 备注"]
    for c, h in enumerate(headers1, 1):
        ws1.cell(row=4, column=c, value=h)
    style_header(ws1, 4, 4)

    items1 = [
        # (序号, 项目, 默认值, 备注)
        (1, "全年营业收入（含税）", 500000.00, "元，取自利润表第1行"),
        (2, "全年营业成本", 200000.00, "元，取自利润表第2行"),
        (3, "全年利润总额", 50000.00, "元，取自利润表第12行"),
        (4, "增值税计税收入（不含税）", 495049.50, "元，小规模纳税人填含税÷1.01"),
        (5, "平均从业人数", 2, "人，全年各季度平均"),
        (6, "平均资产总额", 50.00, "万元"),
        (7, "—— 以下为其他税种参数 ——", "", "——"),
        (8, "注册资本实缴额（累计）", 300000.00, "元，填实际已到位金额（非认缴额），例：注册资本100万已到位30万则填300000"),
        (9, "本年增资额", 0.00, "元，本年新增实收资本，无增资填0"),
        (10, "全年工资总额", 111132.00, "元，全部员工的税前工资年合计"),
        (11, "全年社保公司承担部分", 34560.00, "元，约为工资的25.6%"),
        (12, "上年平均在职职工人数", 2, "人，用于残保金测算"),
        (13, "上年安排残疾人就业人数", 0, "人，需残联审核确认"),
        (14, "上年职工年平均工资", 60000.00, "元/年"),
        (15, "当地社会平均工资（元/年）", 90000.00, "元/年，武汉2024年约8~9万"),
        (16, "—— 季度分摊方式（二选一）——", "", "——"),
        (17, "季度分摊方式", "平均", "填「平均」按4季度均分；填「明细」请在Sheet3填写各季度数据"),
    ]
    for i, (seq, name, val, note) in enumerate(items1):
        r = 5 + i
        ws1.cell(row=r, column=1, value=seq)
        ws1.cell(row=r, column=2, value=name)
        ws1.cell(row=r, column=3, value=val if val != "" else "")
        ws1.cell(row=r, column=4, value=note)
        # 分隔行灰色
        if name.startswith("——"):
            for c in range(1, 5):
                ws1.cell(row=r, column=c).font = Font(name="微软雅黑", size=9, color="9CA3AF", italic=True)
        else:
            style_body_row(ws1, r, 4, is_num=(isinstance(val, (int, float)) and val != ""))

    ws1.column_dimensions["A"].width = 6
    ws1.column_dimensions["B"].width = 30
    ws1.column_dimensions["C"].width = 18
    ws1.column_dimensions["D"].width = 42

    # ============================================================
    #  Sheet 2：员工信息
    # ============================================================
    ws2 = wb.create_sheet("员工信息")
    ws2.merge_cells("A1:H1")
    ws2.cell(row=1, column=1, value="员工工资与专项附加扣除明细").font = Font(name="微软雅黑", bold=True, size=13, color="1E3A8A")
    ws2.cell(row=2, column=1, value="填写说明：每行一名员工，金额均为月均数；员工人数与 Sheet1「平均从业人数」一致").font = hint_font
    ws2.merge_cells("A2:H2")

    headers2 = ["姓名", "税前月工资", "社保缴费基数", "个人社保月实缴", "专项附加扣除合计",
                 "子女教育", "婴幼儿照护", "赡养老人"]
    for c, h in enumerate(headers2, 1):
        ws2.cell(row=4, column=c, value=h)
    style_header(ws2, 4, 8)

    sample_emps = [
        ["张三", 10522, 5000, 522, 5000, 2000, 2000, 1000],
        ["李四", 8000, 5000, 522, 0, 0, 0, 0],
    ]
    for i, emp in enumerate(sample_emps):
        r = 5 + i
        for c, val in enumerate(emp, 1):
            ws2.cell(row=r, column=c, value=val)
        style_body_row(ws2, r, 8, is_num=False)
        # 数字列右对齐
        for c in range(2, 9):
            ws2.cell(row=r, column=c).alignment = num_align

    for c, w in enumerate([10, 14, 14, 16, 18, 12, 12, 12], 1):
        ws2.column_dimensions[get_column_letter(c)].width = w

    # ============================================================
    #  Sheet 3：季度分摊明细（可选）
    # ============================================================
    ws3 = wb.create_sheet("季度分摊明细")
    ws3.merge_cells("A1:E1")
    ws3.cell(row=1, column=1, value="各季度营业收入/成本/利润明细（可选）").font = Font(name="微软雅黑", bold=True, size=13, color="1E3A8A")
    ws3.cell(row=2, column=1, value="填写说明：仅当 Sheet1 季度分摊方式选「明细」时需填写；选「平均」则忽略本表").font = hint_font
    ws3.merge_cells("A2:E2")

    headers3 = ["季度", "营业收入（元）", "营业成本（元）", "利润总额（元）", "增值税计税收入（元）"]
    for c, h in enumerate(headers3, 1):
        ws3.cell(row=4, column=c, value=h)
    style_header(ws3, 4, 5)

    for i, q in enumerate(["Q1", "Q2", "Q3", "Q4"]):
        r = 5 + i
        ws3.cell(row=r, column=1, value=q)
        for c in range(2, 6):
            ws3.cell(row=r, column=c, value=0)
        style_body_row(ws3, r, 5, is_num=True)

    ws3.column_dimensions["A"].width = 8
    for c in range(2, 6):
        ws3.column_dimensions[get_column_letter(c)].width = 18

    # 保存到字节流
    import io
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def gen_annual_report_template_pdf_bytes() -> bytes | None:
    """生成「年报数据导入」PDF 模板（无需填写，用于打印/存档），返回字节流"""
    if not _HAS_FPDF:
        return None
    try:
        from fpdf import FPDF

        pdf = FPDF()
        pdf.set_auto_page_break(auto=True, margin=12)

        # 注册中文字体
        font_path = r"C:\Windows\Fonts\msyh.ttc"
        if not os.path.exists(font_path):
            font_path = r"C:\Windows\Fonts\simsunb.ttf"
        if os.path.exists(font_path):
            pdf.add_font("CJK", "", font_path)
            pdf.add_font("CJK", "B", font_path)
            use_cjk = True
        else:
            use_cjk = False

        def text_cjk(w, h, txt, **kw):
            if use_cjk:
                pdf.set_font("CJK", kw.pop("style", "").replace("B", "B") or "", kw.pop("size", 10))
                pdf.cell(w, h, txt, **kw)
            else:
                safe = txt.encode("ascii", errors="replace").decode("ascii")
                pdf.set_font("Helvetica", kw.pop("style", "").replace("B", "B") or "", kw.pop("size", 10))
                pdf.cell(w, h, safe, **kw)

        def multi_cjk(w, h, txt, **kw):
            if use_cjk:
                pdf.set_font("CJK", kw.pop("style", "").replace("B", "B") or "", kw.pop("size", 10))
                pdf.multi_cell(w, h, txt, **kw)
            else:
                safe = txt.encode("ascii", errors="replace").decode("ascii")
                pdf.set_font("Helvetica", kw.pop("style", "").replace("B", "B") or "", kw.pop("size", 10))
                pdf.multi_cell(w, h, safe, **kw)

        # ── 封面 ──
        pdf.add_page()
        pdf.ln(30)
        if use_cjk:
            pdf.set_font("CJK", "B", 20)
        else:
            pdf.set_font("Helvetica", "B", 20)
        pdf.cell(0, 12, "湖北启贤托儿所有限公司", ln=True, align="C")
        pdf.ln(4)
        if use_cjk:
            pdf.set_font("CJK", "B", 16)
        else:
            pdf.set_font("Helvetica", "B", 16)
        pdf.cell(0, 10, "2025 年度年报数据导入模板", ln=True, align="C")
        pdf.ln(10)
        if use_cjk:
            pdf.set_font("CJK", "", 10)
        else:
            pdf.set_font("Helvetica", "", 10)
        pdf.cell(0, 8, "生成日期：" + datetime.now().strftime("%Y-%m-%d"), ln=True, align="C")
        pdf.cell(0, 8, "说明：本 PDF 为模板参考文件，实际数据请通过 Excel 版导入", ln=True, align="C")
        pdf.cell(0, 8, "Excel 导入路径：系统侧边栏 → 年报导入 → 上传 .xlsx 文件", ln=True, align="C")

        # ── 第1部分：年报汇总 ──
        pdf.add_page()
        if use_cjk:
            pdf.set_font("CJK", "B", 14)
        else:
            pdf.set_font("Helvetica", "B", 14)
        pdf.cell(0, 10, "一、年报汇总数据", ln=True)
        pdf.ln(4)

        col_w = [12, 86, 28, 62]
        headers = ["序号", "项目名称", "数值", "单位/备注"]
        if use_cjk:
            pdf.set_font("CJK", "B", 8)
        else:
            pdf.set_font("Helvetica", "B", 8)
        for i, (h, w) in enumerate(zip(headers, col_w)):
            pdf.cell(w, 7, h, border=1, align="C")
        pdf.ln()

        items = [
            ("1", "全年营业收入（含税）", "500,000.00", "元"),
            ("2", "全年营业成本", "200,000.00", "元"),
            ("3", "全年利润总额", "50,000.00", "元"),
            ("4", "增值税计税收入（不含税）", "495,049.50", "元"),
            ("5", "平均从业人数", "2", "人"),
            ("6", "平均资产总额", "50.00", "万元"),
            ("—", "—— 以下为其他税种参数 ——", "", ""),
            ("7", "注册资本实缴额（累计）", "300,000.00", "元（已到位部分，非认缴全额）"),
            ("8", "本年增资额", "0.00", "元"),
            ("9", "全年工资总额", "111,132.00", "元"),
            ("10", "全年社保公司承担部分", "34,560.00", "元"),
            ("11", "上年平均在职职工人数", "2", "人"),
            ("12", "上年安排残疾人就业人数", "0", "人"),
            ("13", "上年职工年平均工资", "60,000.00", "元/年"),
            ("14", "当地社会平均工资", "90,000.00", "元/年"),
            ("—", "—— 季度分摊方式 ——", "", ""),
            ("15", "季度分摊方式", "平均", "填「平均」或「明细」"),
        ]
        if use_cjk:
            pdf.set_font("CJK", "", 8)
        else:
            pdf.set_font("Helvetica", "", 8)
        for seq, name, val, note in items:
            is_section = seq == "—"
            fs = "CJK" if use_cjk else "Helvetica"
            if is_section:
                pdf.set_font(fs, "", 8)
                pdf.set_text_color(150, 150, 150)
            else:
                pdf.set_font(fs, "", 8)
                pdf.set_text_color(0, 0, 0)
            pdf.cell(col_w[0], 6, seq, border=1, align="C")
            pdf.cell(col_w[1], 6, name, border=1, align="L")
            pdf.cell(col_w[2], 6, val, border=1, align="R")
            pdf.cell(col_w[3], 6, note, border=1, align="L")
            pdf.ln()

        # ── 第2部分：员工信息 ──
        pdf.add_page()
        if use_cjk:
            pdf.set_font("CJK", "B", 14)
        else:
            pdf.set_font("Helvetica", "B", 14)
        pdf.cell(0, 10, "二、员工工资与专项附加扣除明细", ln=True)
        pdf.ln(4)

        emp_col_w = [18, 23, 23, 25, 28, 20, 22, 22]
        emp_headers = ["姓名", "税前月工资", "社保缴费基数", "个人社保月实缴",
                        "专项附加扣除合计", "子女教育", "婴幼儿照护", "赡养老人"]
        if use_cjk:
            pdf.set_font("CJK", "B", 7)
        else:
            pdf.set_font("Helvetica", "B", 7)
        for h, w in zip(emp_headers, emp_col_w):
            pdf.cell(w, 7, h, border=1, align="C")
        pdf.ln()

        sample_emps = [
            ["张三", "10,522", "5,000", "522", "5,000", "2,000", "2,000", "1,000"],
            ["李四", "8,000", "5,000", "522", "0", "0", "0", "0"],
        ]
        if use_cjk:
            pdf.set_font("CJK", "", 7)
        else:
            pdf.set_font("Helvetica", "", 7)
        for emp in sample_emps:
            for val, w in zip(emp, emp_col_w):
                pdf.cell(w, 6, val, border=1, align="R" if val.replace(",", "").replace(".", "").isdigit() else "L")
            pdf.ln()

        # ── 第3部分：季度分摊明细 ──
        pdf.ln(8)
        if use_cjk:
            pdf.set_font("CJK", "B", 14)
        else:
            pdf.set_font("Helvetica", "B", 14)
        pdf.cell(0, 10, "三、季度分摊明细（可选）", ln=True)
        pdf.ln(4)

        q_col_w = [20, 40, 40, 40, 40]
        q_headers = ["季度", "营业收入（元）", "营业成本（元）", "利润总额（元）", "增值税计税收入（元）"]
        if use_cjk:
            pdf.set_font("CJK", "B", 7)
        else:
            pdf.set_font("Helvetica", "B", 7)
        for h, w in zip(q_headers, q_col_w):
            pdf.cell(w, 7, h, border=1, align="C")
        pdf.ln()

        if use_cjk:
            pdf.set_font("CJK", "", 7)
        else:
            pdf.set_font("Helvetica", "", 7)
        for q in ["Q1", "Q2", "Q3", "Q4"]:
            pdf.cell(q_col_w[0], 6, q, border=1, align="C")
            for w in q_col_w[1:]:
                pdf.cell(w, 6, "0", border=1, align="R")
            pdf.ln()

        pdf.ln(6)
        if use_cjk:
            pdf.set_font("CJK", "", 8)
        else:
            pdf.set_font("Helvetica", "", 8)
        pdf.multi_cell(0, 5,
            "注意：本 PDF 为模板参考文件，不可直接导入。请下载 Excel 版本填写数据后上传导入。\n"
            "湖北启贤托儿所有限公司 · 小规模纳税人 · 小型微利企业\n"
            f"模板生成日期：{datetime.now().strftime('%Y-%m-%d %H:%M')}"
        )

        return bytes(pdf.output())
    except Exception as e:
        return None


def parse_annual_report_excel(file_bytes: bytes) -> dict:
    """
    解析年报导入 Excel，返回结构化数据。
    返回格式：
    {
      "summary": { 年报汇总字段 },
      "employees": [ 员工列表 ],
      "quarterly": { "Q1": {...}, "Q2": {...}, "Q3": {...}, "Q4": {...} } or None,
      "warnings": [ 校验警告 ],
    }
    """
    import openpyxl
    import io

    result = {
        "summary": {},
        "employees": [],
        "quarterly": None,
        "warnings": [],
    }

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

    # ── 解析 Sheet1：年报汇总 ──
    if "年报汇总" in wb.sheetnames:
        ws1 = wb["年报汇总"]
        key_map = {
            "全年营业收入（含税）": "annual_revenue",
            "全年营业成本": "annual_cost",
            "全年利润总额": "annual_profit",
            "增值税计税收入（不含税）": "annual_vat_revenue",
            "平均从业人数": "avg_employees",
            "平均资产总额": "avg_assets",
            "注册资本实缴额（累计）": "reg_capital",
            "本年增资额": "capital_increase",
            "全年工资总额": "total_salary",
            "全年社保公司承担部分": "total_si_company",
            "上年平均在职职工人数": "prev_employees",
            "上年安排残疾人就业人数": "prev_disabled",
            "上年职工年平均工资": "prev_avg_salary",
            "当地社会平均工资（元/年）": "local_avg_salary",
            "季度分摊方式": "split_method",
        }
        for row in ws1.iter_rows(min_row=5, values_only=True):
            if row[1] is None:
                continue
            name = str(row[1]).strip()
            if name in key_map:
                val = row[2]
                if val is None:
                    val = 0
                if isinstance(val, str):
                    val = val.strip()
                    if val == "平均":
                        result["summary"][key_map[name]] = "平均"
                    elif val == "明细":
                        result["summary"][key_map[name]] = "明细"
                    else:
                        try:
                            val = float(val.replace(",", "").replace("，", ""))
                        except ValueError:
                            result["warnings"].append(f"「{name}」的值「{val}」无法识别，已跳过")
                            continue
                elif isinstance(val, (int, float)):
                    pass  # 保持原值
                else:
                    result["warnings"].append(f"「{name}」的格式不支持，已跳过")
                    continue
                result["summary"][key_map[name]] = val

    # ── 解析 Sheet2：员工信息 ──
    if "员工信息" in wb.sheetnames:
        ws2 = wb["员工信息"]
        for row in ws2.iter_rows(min_row=5, values_only=True):
            if row[0] is None or str(row[0]).strip() == "":
                continue
            name = str(row[0]).strip()
            if name.startswith("（示例）") or name == "姓名":
                continue
            emp = {
                "name": name,
                "gross_salary": float(row[1] or 0),
                "si_base": float(row[2] or 0),
                "si_personal_actual": float(row[3] or 0),
                "special_deductions": float(row[4] or 0),
                "child_education": float(row[5] or 0),
                "infant_care": float(row[6] or 0),
                "elderly_care": float(row[7] or 0),
            }
            result["employees"].append(emp)

    # ── 解析 Sheet3：季度分摊明细（可选）──
    if "季度分摊明细" in wb.sheetnames:
        ws3 = wb["季度分摊明细"]
        quarterly = {}
        for row in ws3.iter_rows(min_row=5, values_only=True):
            if row[0] is None:
                continue
            q_name = str(row[0]).strip().upper()
            if q_name in ("Q1", "Q2", "Q3", "Q4"):
                quarterly[q_name] = {
                    "revenue": float(row[1] or 0),
                    "cost": float(row[2] or 0),
                    "period_profit": float(row[3] or 0),
                    "vat_revenue": float(row[4] or 0),
                }
        if quarterly:
            result["quarterly"] = quarterly

    wb.close()

    # ── 校验 ──
    s = result["summary"]
    if s.get("annual_revenue", 0) <= 0 and s.get("annual_cost", 0) <= 0:
        result["warnings"].append("营业收入和营业成本均为 0，请确认是否已填入年报数据")
    if s.get("avg_employees", 0) <= 0:
        result["warnings"].append("平均从业人数为 0")
    if len(result["employees"]) == 0:
        result["warnings"].append("未解析到员工信息，请检查「员工信息」Sheet 是否已填写")
    if len(result["employees"]) > 0 and s.get("avg_employees", 0) > 0:
        if len(result["employees"]) != int(s["avg_employees"]):
            result["warnings"].append(
                f"员工人数（{len(result['employees'])}人）与「平均从业人数」（{int(s['avg_employees'])}人）不一致"
            )

    return result


def parse_annual_report_pdf(file_bytes: bytes) -> dict:
    """
    用 pdfplumber 提取 PDF 文本，再用 DeepSeek AI 解析为结构化年报数据。
    返回格式与 parse_annual_report_excel() 一致。
    """
    import pdfplumber
    import io
    import re

    result = {
        "summary": {},
        "employees": [],
        "quarterly": None,
        "warnings": [],
    }

    # ── Step 1：提取 PDF 全文 ──
    full_text_parts = []
    try:
        with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
            for page in pdf.pages:
                text = page.extract_text()
                if text:
                    full_text_parts.append(text)
        full_text = "\n".join(full_text_parts)
    except Exception as e:
        result["warnings"].append(f"PDF 文本提取失败：{e}")
        return result

    if not full_text.strip():
        result["warnings"].append("PDF 中未提取到文字内容，可能为扫描件（图片型 PDF），请使用 Excel 导入")
        return result

    # ── Step 2：DeepSeek AI 解析 ──
    api_key, _ = get_api_key()
    if not api_key:
        result["warnings"].append("未配置 DeepSeek API Key，无法 AI 解析 PDF。请先配置 API Key 或改用 Excel 导入")
        return result

    system_prompt = """你是一个税务数据提取助手。从年报 PDF 文本中提取结构化数据，返回纯 JSON（不要 Markdown 代码块）。

JSON 格式：
{
  "summary": {
    "annual_revenue": 数字（全年营业收入含税，元）,
    "annual_cost": 数字（全年营业成本，元）,
    "annual_profit": 数字（全年利润总额，元）,
    "annual_vat_revenue": 数字（增值税计税收入不含税，元）,
    "avg_employees": 数字（平均从业人数）,
    "avg_assets": 数字（平均资产总额，万元）,
    "reg_capital": 数字（注册资本实缴额累计，元）,
    "capital_increase": 数字（本年增资额，元）,
    "total_salary": 数字（全年工资总额，元）,
    "total_si_company": 数字（全年社保公司承担部分，元）,
    "prev_employees": 数字（上年平均在职职工人数）,
    "prev_disabled": 数字（上年安排残疾人就业人数）,
    "prev_avg_salary": 数字（上年职工年平均工资，元/年）,
    "local_avg_salary": 数字（当地社会平均工资，元/年）
  },
  "employees": [{"name": "姓名", "gross_salary": 月工资, "si_base": 社保基数, "si_personal_actual": 个人社保月实缴, "special_deductions": 专项扣除合计, "child_education": 子女教育, "infant_care": 婴幼儿照护, "elderly_care": 赡养老人}],
  "notes": "补充说明"
}

规则：
- 未找到的字段填 0
- 数字不要带逗号或单位
- 注意区分「全年」和「月均」数据
- 注意区分「含税」和「不含税」
- 注册资本只提取实际到位（实缴）金额，不是认缴金额
- 如果 PDF 只包含年报汇总数据（无员工明细），employees 返回空数组
"""

    prompt = f"""请从以下年报 PDF 文本中提取结构化数据：

===== PDF 文本开始 =====
{full_text[:8000]}
===== PDF 文本结束 =====

请返回 JSON："""

    try:
        ai_response = ask_deepseek(prompt, system_prompt=system_prompt)
    except Exception as e:
        result["warnings"].append(f"AI 解析调用失败：{e}")
        return result

    # ── Step 3：解析 AI 返回的 JSON ──
    # 尝试提取 JSON（处理可能的 Markdown 代码块）
    json_match = re.search(r'\{[\s\S]*\}', ai_response)
    if not json_match:
        result["warnings"].append(f"AI 未返回有效 JSON。原始响应：\n{ai_response[:500]}")
        return result

    try:
        parsed = json.loads(json_match.group())
    except json.JSONDecodeError:
        result["warnings"].append(f"AI 返回的 JSON 解析失败。原始响应：\n{ai_response[:500]}")
        return result

    # ── Step 4：标准化字段 ──
    s = parsed.get("summary", {})
    key_map = {
        "annual_revenue": "annual_revenue",
        "annual_cost": "annual_cost",
        "annual_profit": "annual_profit",
        "annual_vat_revenue": "annual_vat_revenue",
        "avg_employees": "avg_employees",
        "avg_assets": "avg_assets",
        "reg_capital": "reg_capital",
        "capital_increase": "capital_increase",
        "total_salary": "total_salary",
        "total_si_company": "total_si_company",
        "prev_employees": "prev_employees",
        "prev_disabled": "prev_disabled",
        "prev_avg_salary": "prev_avg_salary",
        "local_avg_salary": "local_avg_salary",
    }
    for key, mapped in key_map.items():
        val = s.get(key, 0)
        try:
            result["summary"][mapped] = float(val) if val else 0.0
        except (ValueError, TypeError):
            result["summary"][mapped] = 0.0

    # 员工数据
    raw_emps = parsed.get("employees", [])
    for emp in raw_emps:
        if isinstance(emp, dict) and emp.get("name"):
            result["employees"].append({
                "name": str(emp.get("name", "")),
                "gross_salary": float(emp.get("gross_salary", 0) or 0),
                "si_base": float(emp.get("si_base", 0) or 0),
                "si_personal_actual": float(emp.get("si_personal_actual", 0) or 0),
                "special_deductions": float(emp.get("special_deductions", 0) or 0),
                "child_education": float(emp.get("child_education", 0) or 0),
                "infant_care": float(emp.get("infant_care", 0) or 0),
                "elderly_care": float(emp.get("elderly_care", 0) or 0),
            })

    # 附加说明
    notes = parsed.get("notes", "")
    if notes:
        result["warnings"].append(f"AI 解析备注：{notes}")

    # ── 校验 ──
    s2 = result["summary"]
    if s2.get("annual_revenue", 0) <= 0 and s2.get("annual_cost", 0) <= 0:
        result["warnings"].append("AI 未能从 PDF 中提取到有效的营收/成本数据，请检查 PDF 是否为标准年报格式")
    if len(result["employees"]) == 0:
        result["warnings"].append("PDF 中未提取到员工明细（这是正常的，请手动补充员工数据）")

    # 季度分摊默认平均
    result["summary"]["split_method"] = "平均"

    return result


# ===============================================
#  页面配置
# ===============================================

st.set_page_config(
    page_title="启贤托育AI税务助手",
    page_icon="🧾",
    layout="wide",
)

# ===============================================
#  全局自定义样式
# ===============================================
st.markdown("""
<style>
/* ── 整体风格 ── */
.main .block-container {
    padding-top: 1.5rem;
}

/* ── 指标卡美化 ── */
div[data-testid="stMetric"] {
    background: linear-gradient(135deg, #f8fafc 0%, #f1f5f9 100%);
    border: 1px solid #e2e8f0;
    border-radius: 10px;
    padding: 12px 16px;
    transition: all 0.2s ease;
}
div[data-testid="stMetric"]:hover {
    border-color: #94a3b8;
    box-shadow: 0 2px 8px rgba(0,0,0,0.06);
}
div[data-testid="stMetric"] label {
    font-size: 0.8rem;
    color: #64748b;
}
div[data-testid="stMetric"] div[data-testid="stMetricValue"] {
    font-size: 1.4rem;
    font-weight: 700;
    color: #0f172a;
}

/* ── 表格美化 ── */
div[data-testid="stDataFrame"] {
    border: 1px solid #e2e8f0;
    border-radius: 8px;
    overflow: hidden;
}
div[data-testid="stDataFrame"] th {
    background: #f1f5f9 !important;
    color: #334155 !important;
    font-weight: 600 !important;
    font-size: 0.82rem !important;
}
div[data-testid="stDataFrame"] td {
    font-size: 0.85rem !important;
}

/* ── 主按钮 ── */
button[kind="primary"] {
    background: linear-gradient(135deg, #2563eb 0%, #1d4ed8 100%) !important;
    border: none !important;
    font-weight: 600 !important;
    border-radius: 8px !important;
    letter-spacing: 0.02em;
}
button[kind="primary"]:hover {
    box-shadow: 0 4px 12px rgba(37,99,235,0.35) !important;
    transform: translateY(-1px);
}

/* ── 展开面板 ── */
div[data-testid="stExpander"] {
    border: 1px solid #e2e8f0;
    border-radius: 8px;
    margin-bottom: 0.5rem;
}

/* ── Radio按钮组 ── */
div[data-testid="stRadio"] label {
    font-weight: 500;
}

/* ── 分隔线 ── */
hr {
    margin: 1.2rem 0;
    border-color: #e2e8f0;
}

/* ── Caption文字 ── */
.stCaption {
    color: #64748b;
    font-size: 0.85rem;
}

/* ── Toast消息 ── */
div[data-testid="stToast"] {
    border-radius: 10px !important;
}

/* ── 移动端适配 ── */
@media (max-width: 768px) {
    div[data-testid="stMetric"] div[data-testid="stMetricValue"] {
        font-size: 1.1rem !important;
    }
    .stTabs button {
        font-size: 0.75rem !important;
        padding: 0.4rem 0.6rem !important;
    }
}
</style>
""", unsafe_allow_html=True)

# ===============================================
#  侧边栏
# ===============================================

with st.sidebar:
    st.title("⚙️ 配置")

    # API Key 状态显示
    st.subheader("DeepSeek AI")
    current_key, key_source = get_api_key()

    if key_source in ("secrets", "env"):
        st.success("✅ 已通过配置文件配置 API Key")
        st.caption("如需临时覆盖，可在下方输入")
    elif key_source == "manual":
        st.success("✅ 已手动配置 API Key")
    else:
        st.warning("⚠️ 未配置 API Key")
        st.caption("请在 .streamlit/secrets.toml 中添加 DEEPSEEK_API_KEY=sk-xxx，或在下方输入")

    # 手动输入（可覆盖 .env）
    api_key_manual = st.text_input(
        "手动输入 API Key（可选，覆盖配置文件）",
        value=st.session_state.get("deepseek_api_key_manual", ""),
        type="password",
        help="在 platform.deepseek.com 获取。留空则使用配置文件中的配置。",
        key="deepseek_api_key_manual_input",
    )
    # 同步到 session_state
    if api_key_manual != st.session_state.get("deepseek_api_key_manual", ""):
        st.session_state["deepseek_api_key_manual"] = api_key_manual
        st.rerun()

    # 测试连接按钮
    if st.button("🔍 测试 AI 连接", use_container_width=True):
        test_result = ask_deepseek("请用一句话介绍你自己")
        if test_result.startswith("["):
            st.error(f"连接失败：{test_result}")
        else:
            st.success("✅ AI 连接成功！")
            st.caption(test_result[:100] + "...")

    st.divider()

    # 社保参数说明（可折叠）
    with st.expander("📋 社保参数说明（武汉 2026）", expanded=False):
        st.caption("依据：鄂人社发〔2023〕及武汉医保局最新标准")
        st.markdown("**个人缴纳**")
        st.markdown("- 养老 8%\n- 医疗 2%\n- 失业 **0.3%**\n- 大病医保 **7 元/月**（定额）")
        st.markdown(f"- 合计 ≈ **522 元/月**（基数5000×10.3%+7）")
        st.markdown("**公司缴纳**")
        st.markdown("- 养老 16%\n- 医疗 8.7%（含生育+大病）\n- 失业 **0.7%**\n- 工伤 **0.2%**（一类风险行业）")
        st.markdown(f"- 合计 ≈ **1,280 元/月**（基数5000×25.6%）")
        st.markdown(f"**个税起征点**：{BASIC_DEDUCTION} 元/月")

    st.divider()
    st.subheader("📜 印花税（资金账簿）")
    st.caption("数据来自「🗂️ 年报导入」，此处为只读展示")

    stamp_reg = st.session_state.get("stamp_reg_capital", 0.0)
    stamp_inc = st.session_state.get("stamp_capital_increase", 0.0)

    rc1, rc2 = st.columns(2)
    with rc1:
        st.metric("注册资本实缴（累计）", f"{stamp_reg:,.0f} 元")
    with rc2:
        st.metric("本期增资额", f"{stamp_inc:,.0f} 元")

    if stamp_reg > 0 or stamp_inc > 0:
        from tax_calculator import calc_stamp_duty as _csd
        _preview = _csd(stamp_reg, stamp_inc, 0, 0, 0, 0, 0, True)
        st.caption(f"💡 预估资金账簿印花税：**{_preview['印花税合计（应缴）']:,.2f} 元**（减半后）")
    else:
        st.caption("💡 数据为 0，请先到「🗂️ 年报导入」导入年报数据")

    st.divider()
    st.subheader("♿ 残保金参数")
    def_prev_employees_key = "def_prev_employees"
    if def_prev_employees_key not in st.session_state:
        st.session_state[def_prev_employees_key] = 2
    def_prev_disabled_key = "def_prev_disabled"
    if def_prev_disabled_key not in st.session_state:
        st.session_state[def_prev_disabled_key] = 0
    def_prev_avg_salary_key = "def_prev_avg_salary"
    if def_prev_avg_salary_key not in st.session_state:
        st.session_state[def_prev_avg_salary_key] = 60000.0
    def_local_avg_salary_key = "def_local_avg_salary"
    if def_local_avg_salary_key not in st.session_state:
        st.session_state[def_local_avg_salary_key] = 90000.0
    def_year_key = "def_year"
    if def_year_key not in st.session_state:
        st.session_state[def_year_key] = 2026

    st.session_state[def_prev_employees_key] = st.number_input(
        "上年在职职工人数", min_value=0, value=st.session_state[def_prev_employees_key], step=1,
        help="上年用人单位年平均在职职工人数", key="sb_prev_employees")
    st.session_state[def_prev_disabled_key] = st.number_input(
        "上年安排残疾人就业人数", min_value=0, value=st.session_state[def_prev_disabled_key], step=1,
        help="上年实际安排的残疾人就业人数（需残联审核确认）", key="sb_prev_disabled")
    st.session_state[def_prev_avg_salary_key] = st.number_input(
        "上年职工年平均工资（元）", min_value=0.0, value=st.session_state[def_prev_avg_salary_key], step=1000.0,
        help="上年用人单位在职职工年平均工资", key="sb_prev_avg_salary")
    st.session_state[def_local_avg_salary_key] = st.number_input(
        "当地社会平均工资（元/年）", min_value=0.0, value=st.session_state[def_local_avg_salary_key], step=1000.0,
        help="武汉2024年社平工资约8~9万/年，用于2倍封顶", key="sb_local_avg_salary")
    st.session_state[def_year_key] = st.number_input(
        "申报年度", min_value=2024, max_value=2030, value=st.session_state[def_year_key], step=1, key="sb_year")

    st.divider()
    st.caption("启贤托育AI税务助手 v1.6")
    st.caption("仅供参考，申报前请核实")

# ===============================================
#  主界面
# ===============================================

st.title("💰 启贤托育AI税务助手")
st.caption("湖北启贤托儿所有限公司 · 一站式企业税务计算与AI申报指引")

tab1, tab2, tab3, tab4, tab5, tab6, tab7 = st.tabs(["🗂️ 年报导入", "📊 季度申报", "💰 个税计算", "🏦 税款缴纳清单", "♿ 残保金", "📄 申报说明", "📖 申报指南"])

# ---- Tab1：手动录入 ----

with tab3:
    st.header("💰 个税计算")

    # ── 月份选择 + 模式选择（共享）──
    month_col1, month_col2, month_col3 = st.columns([1, 1, 2])
    with month_col1:
        salary_month = st.selectbox(
            "📅 工资所属月份",
            options=list(range(1, 13)),
            format_func=lambda m: f"{m}月",
            index=datetime.now().month - 1,
            key="salary_month",
        )
    with month_col2:
        salary_year = st.selectbox(
            "年份",
            options=list(range(2024, 2031)),
            index=2 if datetime.now().year == 2026 else 0,
            key="salary_year",
        )
    with month_col3:
        from calendar import monthrange
        _, last_day = monthrange(salary_year, salary_month)
        if salary_month == 12:
            deadline_month, deadline_year = 1, salary_year + 1
        else:
            deadline_month, deadline_year = salary_month + 1, salary_year
        deadline_day = 15
        from datetime import date as dt_date
        deadline_date = dt_date(deadline_year, deadline_month, deadline_day)
        while deadline_date.weekday() >= 5:
            deadline_date = dt_date(deadline_date.year, deadline_date.month, deadline_date.day + 1)
        today = dt_date.today()
        days_left = (deadline_date - today).days
        if days_left < 0:
            deadline_badge = "⚠️ 已过期"
            deadline_color = "red"
        elif days_left <= 3:
            deadline_badge = f"🔴 仅剩 {days_left} 天"
            deadline_color = "red"
        elif days_left <= 7:
            deadline_badge = f"🟡 剩余 {days_left} 天"
            deadline_color = "orange"
        else:
            deadline_badge = f"🟢 剩余 {days_left} 天"
            deadline_color = "green"
        st.markdown(f"""
        <div style="border:1px solid #e0e0e0; border-radius:8px; padding:10px 14px; margin-top:4px;">
            <span style="font-size:13px; color:#666;">📮 个税申报截止日</span><br>
            <span style="font-size:20px; font-weight:700; color:#{deadline_color.replace('red','d32f2f').replace('orange','f57c00').replace('green','388e3c')};">
                {deadline_date.strftime('%Y年%m月%d日')}
            </span>
            <span style="font-size:13px; color:#{deadline_color.replace('red','d32f2f').replace('orange','f57c00').replace('green','388e3c')}; margin-left:8px;">
                {deadline_badge}
            </span>
            <br><span style="font-size:11px; color:#999;">自然人电子税务局（扣缴端）· 次月15日前</span>
        </div>
        """, unsafe_allow_html=True)

    st.divider()

    # ── 模式选择 ──
    calc_mode = st.radio(
        "计算模式",
        ["🧑 单人计算", "📋 批量计算"],
        horizontal=True,
        key="calc_mode",
    )

    salary_period = f"{salary_year}年{salary_month}月"
    deadline_period = f"{deadline_year}年{deadline_month}月{deadline_date.day}日"

    if calc_mode == "🧑 单人计算":
        # 初始化持久化存储
        EMP_DATA_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "员工数据_草稿.json")
        if "employees_saved" not in st.session_state:
            if os.path.exists(EMP_DATA_FILE):
                try:
                    with open(EMP_DATA_FILE, "r", encoding="utf-8") as f:
                        st.session_state["employees_saved"] = json.load(f)
                except Exception:
                    st.session_state["employees_saved"] = []
            else:
                st.session_state["employees_saved"] = []

        saved_emps = st.session_state["employees_saved"]
        saved_count = len(saved_emps)

        top_col1, top_col2, top_col3 = st.columns([2, 1, 1])
        with top_col1:
            default_n = max(saved_count, 1)
            num_emp = st.number_input("员工人数", min_value=1, max_value=20, value=default_n, step=1, key="num_emp_tab1")
        with top_col2:
            if st.button("💾 保存草稿", use_container_width=True):
                draft = []
                for i in range(num_emp):
                    draft.append({
                        "name": st.session_state.get(f"name_{i}", f"员工{i+1}"),
                        "gross_salary": st.session_state.get(f"salary_{i}", 10522.0 if i == 0 else 8000.0),
                        "si_base": st.session_state.get(f"si_base_{i}", 5000.0),
                        "si_personal_actual": st.session_state.get(f"si_personal_{i}", float(SOCIAL_INSURANCE_ACTUAL)),
                        "special_deductions": st.session_state.get(f"special_{i}", 5000.0 if i == 0 else 0.0),
                        "child_education": st.session_state.get(f"child_{i}", 2000.0 if i == 0 else 0.0),
                        "infant_care": st.session_state.get(f"infant_{i}", 2000.0 if i == 0 else 0.0),
                        "elderly_care": st.session_state.get(f"elderly_{i}", 1000.0 if i == 0 else 0.0),
                    })
                with open(EMP_DATA_FILE, "w", encoding="utf-8") as f:
                    json.dump(draft, f, ensure_ascii=False, indent=2)
                st.session_state["employees_saved"] = draft
                st.toast("✅ 草稿已保存", icon="💾")
        with top_col3:
            if saved_count > 0:
                if st.button("📂 加载草稿", use_container_width=True):
                    st.toast(f"✅ 已加载 {saved_count} 名员工", icon="📂")
                    st.rerun()
            else:
                st.button("📂 无草稿", disabled=True, use_container_width=True)

        if saved_count > 0:
            try:
                mtime = datetime.fromtimestamp(os.path.getmtime(EMP_DATA_FILE))
                st.caption(f"💡 上次保存：{mtime.strftime('%m-%d %H:%M')} · {saved_count} 名员工 · 切换「员工人数」后点「加载草稿」恢复")
            except Exception:
                pass

        employees_data = []

        for i in range(num_emp):
            st.divider()
            st.subheader(f"员工 {i+1}")

            default_name = saved_emps[i]["name"] if i < saved_count else f"员工{i+1}"
            default_salary = saved_emps[i]["gross_salary"] if i < saved_count else (10522.0 if i == 0 else 8000.0)
            default_si_base = saved_emps[i]["si_base"] if i < saved_count else 5000.0
            default_si_personal = saved_emps[i]["si_personal_actual"] if i < saved_count else float(SOCIAL_INSURANCE_ACTUAL)
            default_special = saved_emps[i]["special_deductions"] if i < saved_count else (5000.0 if i == 0 else 0.0)
            default_child = saved_emps[i]["child_education"] if i < saved_count else (2000.0 if i == 0 else 0.0)
            default_infant = saved_emps[i]["infant_care"] if i < saved_count else (2000.0 if i == 0 else 0.0)
            default_elderly = saved_emps[i]["elderly_care"] if i < saved_count else (1000.0 if i == 0 else 0.0)

            c1, c2 = st.columns(2)
            with c1:
                name = st.text_input("姓名", value=default_name, key=f"name_{i}")
                salary = st.number_input("税前工资（元）", min_value=0.0, value=default_salary, step=100.0, key=f"salary_{i}")
                si_base = st.number_input("社保缴费基数（元）", min_value=0.0, value=default_si_base, step=100.0, key=f"si_base_{i}")
            with c2:
                si_personal = st.number_input("个人社保实缴（元）", min_value=0.0, value=default_si_personal, step=10.0, key=f"si_personal_{i}")
                special = st.number_input("专项附加扣除合计（元）", min_value=0.0, value=default_special, step=500.0, key=f"special_{i}")
                st.markdown("**专项附加扣除明细**")
                cc1, cc2, cc3 = st.columns(3)
                with cc1:
                    child = st.number_input("子女教育", min_value=0.0, value=default_child, step=500.0, key=f"child_{i}")
                with cc2:
                    infant = st.number_input("婴幼儿照护", min_value=0.0, value=default_infant, step=500.0, key=f"infant_{i}")
                with cc3:
                    elderly = st.number_input("赡养老人", min_value=0.0, value=default_elderly, step=500.0, key=f"elderly_{i}")

            employees_data.append({
                "name": name,
                "gross_salary": salary,
                "si_base": si_base,
                "si_personal_actual": si_personal,
                "special_deductions": special,
                "child_education": child,
                "infant_care": infant,
                "elderly_care": elderly,
            })

        if st.button(f"🚀 开始计算 {salary_period} 个税", use_container_width=True, type="primary"):
            validation_warnings = []
            for emp in employees_data:
                nm = emp["name"]
                gs = emp["gross_salary"]
                sb = emp["si_base"]
                sp = emp["si_personal_actual"]
                sd = emp["special_deductions"]
                ce = emp["child_education"]
                ic = emp["infant_care"]
                ec = emp["elderly_care"]
                if gs <= 0:
                    validation_warnings.append(f"⚠️ {nm}：税前工资为 0，请确认是否遗漏")
                if gs > 0 and sb <= 0:
                    validation_warnings.append(f"⚠️ {nm}：社保缴费基数为 0，请确认")
                if gs > 0 and sb > 0 and sb < gs * 0.4:
                    validation_warnings.append(f"💡 {nm}：社保基数（{sb:.0f}）偏低，通常为工资的 60%~300%")
                if sd > 0 and ce + ic + ec != sd:
                    if abs(ce + ic + ec - sd) > 1:
                        validation_warnings.append(f"💡 {nm}：专项附加扣除合计（{sd:.0f}）与明细之和（{ce+ic+ec:.0f}）不一致")
                if ce > 2000:
                    validation_warnings.append(f"💡 {nm}：子女教育扣除 {ce:.0f} 元超出标准（2000元/人），请核实")
                if ic > 2000:
                    validation_warnings.append(f"💡 {nm}：婴幼儿照护扣除 {ic:.0f} 元超出标准（2000元/人），请核实")
                if ec > 3000:
                    validation_warnings.append(f"💡 {nm}：赡养老人扣除 {ec:.0f} 元超出标准（最高3000元），请核实")

            if validation_warnings:
                with st.expander(f"🔍 数据校验提示（{len(validation_warnings)} 条）", expanded=True):
                    for w in validation_warnings:
                        if w.startswith("⚠️"):
                            st.warning(w)
                        else:
                            st.info(w)

            results = []
            for emp in employees_data:
                r = calc_one_employee(
                    emp["name"], emp["gross_salary"], emp["si_base"],
                    emp["si_personal_actual"], emp["special_deductions"],
                    emp["child_education"], emp["infant_care"], emp["elderly_care"],
                )
                results.append(r)

            st.session_state["results"] = results
            st.session_state["_calc_period"] = salary_period
            st.session_state["_deadline_period"] = deadline_period
            st.success(f"✅ {salary_period} 个税计算完成！")
            st.rerun()

    else:
        # ── 批量计算个税 ──
        col_info, col_template = st.columns([2, 1])
        with col_info:
            st.info(
                "请上传 CSV 或 Excel 文件，需包含以下列：\n"
                "姓名, 税前工资, 社保基数, 个人社保实缴, 专项附加扣除, "
                "子女教育, 婴幼儿照护, 赡养老人"
            )
        with col_template:
            template_csv_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "申报底稿模板.csv")
            if os.path.exists(template_csv_path):
                with open(template_csv_path, "rb") as f:
                    csv_bytes = f.read()
                st.download_button(
                    label="📥 下载上传模板（CSV）",
                    data=csv_bytes, file_name="申报底稿模板.csv",
                    mime="text/csv", use_container_width=True,
                )
            template_xlsx_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "申报底稿模板.xlsx")
            if os.path.exists(template_xlsx_path):
                with open(template_xlsx_path, "rb") as f:
                    xlsx_bytes = f.read()
                st.download_button(
                    label="📥 下载上传模板（Excel）",
                    data=xlsx_bytes, file_name="申报底稿模板.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )

        st.divider()
        uploaded = st.file_uploader("选择文件上传", type=["csv", "xlsx", "xls"], key="uploader")

        if uploaded is not None:
            try:
                file_id = getattr(uploaded, "file_id", id(uploaded))
                if st.session_state.get("_uploaded_file_id") != file_id:
                    if uploaded.name.endswith(".csv"):
                        try:
                            df_up = pd.read_csv(uploaded, encoding="utf-8-sig")
                        except Exception:
                            uploaded.seek(0)
                            df_up = pd.read_csv(uploaded, encoding="gbk")
                    else:
                        df_up = pd.read_excel(uploaded)
                    st.session_state["_uploaded_df"] = df_up
                    st.session_state["_uploaded_file_id"] = file_id
                else:
                    df_up = st.session_state["_uploaded_df"]

                st.write("文件预览：")
                st.dataframe(df_up.head(), use_container_width=True)

                if st.button("🚀 导入并计算", key="btn_upload"):
                    results = []
                    for _, row in df_up.iterrows():
                        r = calc_one_employee(
                            str(row.get("姓名", "员工")),
                            float(row.get("税前工资") or 0),
                            float(row.get("社保基数") or 5000),
                            float(row.get("个人社保实缴") or SOCIAL_INSURANCE_ACTUAL),
                            float(row.get("专项附加扣除") or 0),
                            float(row.get("子女教育") or 0),
                            float(row.get("婴幼儿照护") or 0),
                            float(row.get("赡养老人") or 0),
                        )
                        results.append(r)

                    st.session_state["results"] = results
                    st.session_state["_calc_period"] = salary_period
                    st.session_state["_deadline_period"] = deadline_period
                    st.success(f"✅ 批量计算完成！共 {len(results)} 名员工")
                    st.rerun()
            except Exception as e:
                st.error(f"文件读取失败：{e}")

    # ═══ 共享结果区（单人 + 批量共用）═══
    if "results" in st.session_state and st.session_state["results"]:
        results = st.session_state["results"]
        calc_period = st.session_state.get("_calc_period", salary_period)
        deadline_str = st.session_state.get("_deadline_period", deadline_period)

        df = pd.DataFrame(results)
        st.divider()
        st.subheader(f"📊 计算结果 — {calc_period}")
        numeric_cols = df.select_dtypes(include=["float64", "int64"]).columns
        st.dataframe(df.style.format("{:.2f}", subset=numeric_cols), use_container_width=True)

        st.subheader("📈 汇总")
        m1, m2, m3, m4 = st.columns(4)
        m1.metric("工资总额", f"{df['税前工资'].sum():.2f} 元")
        m2.metric("个税总额", f"{df['应纳税额'].sum():.2f} 元")
        m3.metric("实发总额", f"{df['实发工资'].sum():.2f} 元")
        m4.metric("公司总成本", f"{df['公司用人总成本'].sum():.2f} 元")

        deadline_date_obj = dt_date(deadline_year, deadline_month, deadline_date.day)
        days_left = (deadline_date_obj - dt_date.today()).days
        if days_left >= 0:
            if days_left <= 3:
                st.error(f"⚠️ 申报截止日：**{deadline_str}** — 仅剩 **{days_left}** 天！请尽快在自然人电子税务局（扣缴端）完成申报。")
            elif days_left <= 7:
                st.warning(f"⏰ 申报截止日：**{deadline_str}** — 剩余 **{days_left}** 天。")
            else:
                st.info(f"📮 申报截止日：**{deadline_str}** — 剩余 **{days_left}** 天，请在截止前完成申报。")
        else:
            st.error(f"🚨 申报截止日 **{deadline_str}** 已过期 **{abs(days_left)}** 天！如有逾期，请尽快补报并联系税务机关说明情况。")

        csv_data = df.to_csv(index=False, encoding="utf-8-sig")
        dl_col1, dl_col2 = st.columns(2)
        with dl_col1:
            st.download_button(
                label="📥 下载申报底稿（CSV）",
                data=csv_data,
                file_name=f"申报底稿_{salary_year}{salary_month:02d}.csv",
                mime="text/csv", use_container_width=True,
            )
        with dl_col2:
            summary_lines = [
                f"工资总额: {df['税前工资'].sum():.2f} 元",
                f"个税总额: {df['应纳税额'].sum():.2f} 元",
                f"实发总额: {df['实发工资'].sum():.2f} 元",
                f"公司总成本: {df['公司用人总成本'].sum():.2f} 元",
            ]
            pdf_bytes = make_pdf_with_dataframe("个税申报底稿", df, summary_lines, "")
            if pdf_bytes:
                st.download_button(
                    label="📥 下载申报底稿（PDF）",
                    data=pdf_bytes,
                    file_name=f"申报底稿_{salary_year}{salary_month:02d}.pdf",
                    mime="application/pdf", use_container_width=True,
                )

        with st.expander("🔎 工资数据校验（银行流水 vs 个税申报 vs 年报）", expanded=False):
            st.caption("上传银行流水和/或个税申报记录，与系统录入工资交叉比对")
            val_col1, val_col2 = st.columns(2)
            with val_col1:
                bank_file_val = st.file_uploader("上传银行流水（用于校验工资支出）", type=["csv", "xlsx", "xls"], key="salary_val_bank")
            with val_col2:
                tax_file_val = st.file_uploader("上传个税申报记录（用于校验代扣代缴）", type=["csv", "xlsx", "xls"], key="salary_val_tax")
            annual_salary_input = st.number_input("年报中的「全年工资总额」（元，选填）", min_value=0.0, value=0.0, step=1000.0, key="salary_val_annual")

            if st.button("🔍 执行校验", key="salary_val_btn"):
                warnings = []
                total_salary = df["税前工资"].sum()
                if bank_file_val is not None:
                    try:
                        if bank_file_val.name.endswith(".csv"):
                            try:
                                df_bank = pd.read_csv(bank_file_val, encoding="utf-8-sig")
                            except Exception:
                                bank_file_val.seek(0)
                                df_bank = pd.read_csv(bank_file_val, encoding="gbk")
                        else:
                            df_bank = pd.read_excel(bank_file_val)
                        bank_salary = 0.0
                        for col in df_bank.columns:
                            if any(k in str(col) for k in ["工资", "薪酬", "salary", "payroll"]):
                                bank_salary += df_bank[col].sum()
                        if bank_salary > 0:
                            diff = abs(total_salary - bank_salary)
                            if diff < 1:
                                warnings.append("✅ 银行流水工资支出与系统录入一致")
                            else:
                                warnings.append(f"⚠️ 差异 {diff:.0f} 元（系统 {total_salary:.0f} vs 银行 {bank_salary:.0f}），请核实")
                    except Exception:
                        warnings.append("⚠️ 银行流水解析失败，请确认文件格式")
                if tax_file_val is not None:
                    try:
                        if tax_file_val.name.endswith(".csv"):
                            try:
                                df_tax = pd.read_csv(tax_file_val, encoding="utf-8-sig")
                            except Exception:
                                tax_file_val.seek(0)
                                df_tax = pd.read_csv(tax_file_val, encoding="gbk")
                        else:
                            df_tax = pd.read_excel(tax_file_val)
                        tax_salary = 0.0
                        for col in df_tax.columns:
                            if any(k in str(col) for k in ["收入", "工资", "income"]):
                                tax_salary += df_tax[col].sum()
                        if tax_salary > 0:
                            diff = abs(total_salary - tax_salary)
                            if diff < 1:
                                warnings.append("✅ 个税申报收入额与系统录入一致")
                            else:
                                warnings.append(f"⚠️ 差异 {diff:.0f} 元（系统 {total_salary:.0f} vs 申报 {tax_salary:.0f}），请核实")
                    except Exception:
                        warnings.append("⚠️ 个税申报记录解析失败，请确认文件格式")
                if annual_salary_input > 0:
                    diff = abs(total_salary - annual_salary_input)
                    if diff < 1:
                        warnings.append("✅ 系统录入与年报工资总额一致")
                    else:
                        warnings.append(f"⚠️ 差异 {diff:.0f} 元（系统 {total_salary:.0f} vs 年报 {annual_salary_input:.0f}），请核实")
                if not warnings:
                    warnings.append("💡 未上传校验文件，无法比对")
                if warnings:
                    with st.expander(f"📋 校验说明（{len(warnings)} 条）", expanded=True):
                        for w in warnings:
                            if "✅" in w:
                                st.success(w)
                            elif "⚠️" in w or "差" in w:
                                st.warning(w)
                            else:
                                st.info(w)



# ---- Tab6：申报说明 ----
with tab6:
    st.header("AI 申报说明")

    if "results" not in st.session_state:
        st.info("💡 请先在「💰 个税计算」页面录入员工信息并点击「开始计算」，再返回本页查看 AI 生成的申报说明。")
        st.caption("AI 将根据您的工资数据自动生成个税和社保的申报操作说明。")
    else:
        results = st.session_state["results"]
        now_str = datetime.now().strftime("%Y年%m月")

        # 个税说明
        st.subheader("📄 个税申报说明")
        with st.spinner("AI 正在生成个税申报说明..."):
            tax_text = generate_tax_report_ai(results)
        st.text_area("个税申报说明", tax_text, height=400, key="tax_area")

        # 社保说明
        st.subheader("📄 社保申报说明")
        with st.spinner("AI 正在生成社保申报说明..."):
            social_text = generate_social_report_ai(results)
        st.text_area("社保申报说明", social_text, height=400, key="social_area")

        # 下载 TXT + PDF
        full_text = tax_text + "\n\n" + "=" * 50 + "\n\n" + social_text
        dl_a1, dl_a2 = st.columns(2)
        with dl_a1:
            st.download_button(
                label="📥 下载申报说明（TXT）",
                data=full_text,
                file_name=f"申报说明_{datetime.now().strftime('%Y%m')}.txt",
                mime="text/plain",
                use_container_width=True,
            )
        with dl_a2:
            pdf_bytes = make_pdf(
                f"申报说明 - {datetime.now().strftime('%Y年%m月')}",
                full_text.split("\n"),
                ""
            )
            if pdf_bytes:
                st.download_button(
                    label="📥 下载申报说明（PDF）",
                    data=pdf_bytes,
                    file_name=f"申报说明_{datetime.now().strftime('%Y%m')}.pdf",
                    mime="application/pdf",
                    use_container_width=True,
                )

# ===============================================
#  季度申报数据持久化
# ===============================================

QUARTER_DATA_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "季度申报数据.json")


def load_quarter_data(year: int) -> dict:
    """加载某年度的季度申报数据"""
    if not os.path.exists(QUARTER_DATA_FILE):
        return {}
    try:
        with open(QUARTER_DATA_FILE, "r", encoding="utf-8") as f:
            all_data = json.load(f)
            return all_data.get(str(year), {})
    except Exception:
        return {}


def save_quarter_data(year: int, quarter: int, data: dict):
    """保存季度申报数据"""
    if os.path.exists(QUARTER_DATA_FILE):
        with open(QUARTER_DATA_FILE, "r", encoding="utf-8") as f:
            all_data = json.load(f)
    else:
        all_data = {}

    if str(year) not in all_data:
        all_data[str(year)] = {}

    all_data[str(year)][str(quarter)] = data
    all_data[str(year)]["_updated"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    with open(QUARTER_DATA_FILE, "w", encoding="utf-8") as f:
        json.dump(all_data, f, ensure_ascii=False, indent=2)


def get_ytd_values(year: int, quarter: int) -> dict:
    """获取本年累计值（基于已保存的上季度数据）"""
    data = load_quarter_data(year)
    ytd_revenue = 0.0
    ytd_cost = 0.0
    ytd_profit = 0.0

    for q in range(1, quarter):
        if str(q) in data:
            q_data = data[str(q)]
            ytd_revenue += q_data.get("revenue", 0)
            ytd_cost += q_data.get("cost", 0)
            ytd_profit += q_data.get("period_profit", 0)

    return {
        "ytd_revenue": ytd_revenue,
        "ytd_cost": ytd_cost,
        "ytd_profit": ytd_profit,
    }


# ===============================================
#  Tab4：季度企业所得税申报
# ===============================================

# ---- Tab4：季度企业所得税申报 ----
with tab2:
    st.header("📊 企业所得税季度预缴申报")

    st.info(
        "小型微利企业优惠税率 5%（2024-2027年政策）。\n"
        "系统会自动加载上季度数据，计算本年累计值。"
    )

    # ========== 季度选择 ==========
    col_q, col_y = st.columns([1, 1])
    with col_q:
        quarter = st.selectbox("申报季度", [1, 2, 3, 4], index=min(datetime.now().month // 3, 3))
    with col_y:
        year = st.number_input("年度", min_value=2024, max_value=2030, value=datetime.now().year)

    # ========== 加载上季度数据 ==========
    ytd = get_ytd_values(year, quarter)
    prev_saved = load_quarter_data(year)

    if quarter > 1 and str(quarter - 1) in prev_saved:
        st.success(
            f"✅ 已加载 Q{quarter-1} 数据："
            f"累计收入 {ytd['ytd_revenue']:.2f} 元，"
            f"累计利润 {ytd['ytd_profit']:.2f} 元"
        )

    # ========== 自动加载当前季度已保存数据 ==========
    current_q_key = str(quarter)
    current_saved = prev_saved.get(current_q_key, None) if prev_saved else None

    if current_saved:
        q_source = current_saved.get("_source", "手动录入")
        st.info(
            f"📂 已检测到 **{year}年Q{quarter}** 已有申报数据（来源：{q_source}），"
            f"已自动填入下方表单。如需修改可直接编辑。",
            icon="📋"
        )
        # 自动填入表单
        st.session_state["auto_revenue"] = float(current_saved.get("revenue", 0))
        st.session_state["auto_cost"] = float(current_saved.get("cost", 0))
        st.session_state["auto_profit"] = float(current_saved.get("period_profit", 0))
        st.session_state["auto_vat_revenue"] = float(current_saved.get("vat_revenue",
            current_saved.get("revenue", 0)))
        st.session_state["auto_num_employees"] = int(current_saved.get("num_employees", 1))
        st.session_state["auto_total_assets"] = float(current_saved.get("total_assets", 0))

    # 年度切换时清除旧的自动填入（避免2025数据残留在2026）
    prev_auto_year = st.session_state.get("_last_auto_year", None)
    if prev_auto_year != year:
        for k in ["auto_revenue", "auto_cost", "auto_profit", "auto_vat_revenue",
                  "auto_num_employees", "auto_total_assets", "vat_data", "corp_tax_result"]:
            st.session_state.pop(k, None)
    st.session_state["_last_auto_year"] = year

    st.divider()

    # ========== 银行流水导入区域 ==========
    with st.expander("📥 导入银行流水（自动填表）", expanded=False):
        st.caption("支持民生银行、建设银行等 CSV/Excel 流水文件，自动分类并填入下方表单")

        bank_file = st.file_uploader(
            "上传银行流水文件（可多次上传不同银行）",
            type=["csv", "xlsx", "xls"],
            key="bank_uploader",
            accept_multiple_files=True,
        )

        if bank_file:
            try:
                all_txns = []
                for bf in bank_file:
                    # 读取文件
                    if bf.name.endswith(".csv"):
                        try:
                            df_bank = pd.read_csv(bf, encoding="utf-8-sig")
                        except Exception:
                            bf.seek(0)
                            df_bank = pd.read_csv(bf, encoding="gbk")
                    else:
                        df_bank = pd.read_excel(bf)

                    # ── 民生银行「活期账户明细」格式检测 ──
                    # 前17行为元数据（账户信息/查询参数），第18行为真实表头
                    if not bf.name.endswith(".csv"):
                        first_col = str(df_bank.columns[0]) if len(df_bank.columns) > 0 else ""
                        if "账户名称" in first_col:
                            bf.seek(0)
                            df_bank = pd.read_excel(bf, skiprows=17, header=0)
                            # 跳过汇总行（借方累计/贷方累计）和空行
                            df_bank = df_bank[df_bank.iloc[:, 0].notna()].copy()
                            df_bank = df_bank[~df_bank.iloc[:, 0].astype(str).str.contains("累计|笔数", na=False)].copy()

                    # 统一列名（常见银行格式兼容）
                    col_map = {}
                    for col in df_bank.columns:
                        col_lower = str(col).strip().lower()
                        if any(k in col_lower for k in ["日期", "date", "交易日期", "记账日期", "交易时间", "时间"]):
                            col_map[col] = "交易日期"
                        elif any(k in col_lower for k in ["摘要", "备注", "用途", "description", "摘要说明", "附言"]):
                            col_map[col] = "摘要"
                        elif any(k in col_lower for k in ["收入", "贷方", "存款", "credit", "存入"]):
                            col_map[col] = "收入金额"
                        elif any(k in col_lower for k in ["支出", "借方", "取款", "debit", "转出"]):
                            col_map[col] = "支出金额"
                        elif any(k in col_lower for k in ["金额", "发生额", "transaction"]):
                            col_map[col] = "金额"
                        elif any(k in col_lower for k in ["余额", "balance"]):
                            col_map[col] = "余额"
                        elif any(k in col_lower for k in ["借贷", "收支方向", "类型"]):
                            col_map[col] = "借贷标识"

                    df_bank = df_bank.rename(columns=col_map)

                    # 如果没有明确的收入/支出列，尝试从"金额"+"借贷标识"推断
                    if "金额" in df_bank.columns and "借贷标识" in df_bank.columns:
                        for _, row in df_bank.iterrows():
                            amount = abs(float(row.get("金额") or 0))
                            flag = str(row.get("借贷标识", "")).strip()
                            txn = {
                                "银行": bf.name,
                                "日期": row.get("交易日期", ""),
                                "摘要": row.get("摘要", ""),
                                "收入金额": amount if flag in ["贷", "收入", "存入", "CREDIT"] else 0,
                                "支出金额": amount if flag in ["借", "支出", "转出", "DEBIT"] else 0,
                            }
                            all_txns.append(txn)
                    else:
                        # 直接取收入/支出列
                        for _, row in df_bank.iterrows():
                            all_txns.append({
                                "银行": bf.name,
                                "日期": row.get("交易日期", row.get("日期", "")),
                                "摘要": row.get("摘要", ""),
                                "收入金额": float(row.get("收入金额", 0) or 0),
                                "支出金额": float(row.get("支出金额", 0) or 0),
                            })

                df_txns = pd.DataFrame(all_txns)
                st.success(f"✅ 成功读取 {len(df_txns)} 条交易记录")

                # 使用优化后的分类函数（遵循小企业会计准则）
                df_txns["自动分类"] = df_txns["摘要"].apply(
                    lambda x: classify_bank_transaction(x)["category"]
                )
                df_txns["会计科目"] = df_txns["摘要"].apply(
                    lambda x: classify_bank_transaction(x)["account"]
                )
                df_txns["利润表项目"] = df_txns["摘要"].apply(
                    lambda x: classify_bank_transaction(x)["pl_item"]
                )
                
                st.dataframe(
                    df_txns[["日期", "摘要", "收入金额", "支出金额", "自动分类", "利润表项目"]].head(10), 
                    use_container_width=True
                )

                st.subheader("请确认交易分类（可手动修改）")
                edited_df = st.data_editor(
                    df_txns[["日期", "摘要", "收入金额", "支出金额", "自动分类", "利润表项目"]],
                    use_container_width=True,
                    num_rows="dynamic",
                    key="txn_editor",
                )

                # 生成利润表预览 — 会小企02表格式
                st.subheader("📊 自动生成利润表（会小企02表）")
                profit_data = generate_profit_statement(edited_df)
                
                profit_df = pd.DataFrame({
                    "项目": [
                        "一、营业收入",
                        "减：营业成本",
                        "    税金及附加",
                        "    销售费用",
                        "    管理费用",
                        "    财务费用",
                        "加：投资收益",
                        "二、营业利润",
                        "加：营业外收入",
                        "减：营业外支出",
                        "三、利润总额",
                        "减：所得税费用",
                        "四、净利润",
                    ],
                    "行次": [1, 2, 3, 11, 14, 18, 20, 21, 22, 24, 30, 31, 32],
                    "本月金额": [
                        f"{profit_data['营业收入']:,.2f}",
                        f"{profit_data['营业成本']:,.2f}",
                        f"{profit_data['税金及附加']:,.2f}",
                        f"{profit_data['销售费用']:,.2f}",
                        f"{profit_data['管理费用']:,.2f}",
                        f"{profit_data['财务费用']:,.2f}",
                        f"{profit_data['投资收益']:,.2f}",
                        f"{profit_data['营业利润']:,.2f}",
                        f"{profit_data['营业外收入']:,.2f}",
                        f"{profit_data['营业外支出']:,.2f}",
                        f"{profit_data['利润总额']:,.2f}",
                        f"{profit_data['所得税费用']:,.2f}",
                        f"{profit_data['净利润']:,.2f}",
                    ],
                    "本年累计金额": [
                        f"{profit_data['营业收入']:,.2f}",
                        f"{profit_data['营业成本']:,.2f}",
                        f"{profit_data['税金及附加']:,.2f}",
                        f"{profit_data['销售费用']:,.2f}",
                        f"{profit_data['管理费用']:,.2f}",
                        f"{profit_data['财务费用']:,.2f}",
                        f"{profit_data['投资收益']:,.2f}",
                        f"{profit_data['营业利润']:,.2f}",
                        f"{profit_data['营业外收入']:,.2f}",
                        f"{profit_data['营业外支出']:,.2f}",
                        f"{profit_data['利润总额']:,.2f}",
                        f"{profit_data['所得税费用']:,.2f}",
                        f"{profit_data['净利润']:,.2f}",
                    ],
                })
                st.dataframe(profit_df, use_container_width=True, hide_index=True)
                st.caption("💡 行次对应 会小企02表（小企业会计准则利润表），本月金额=本年累计金额（季度申报）")

                # ═══ 资产负债表项目（不进入利润表）═══
                bs_items = []
                if profit_data.get("其他应付款_股东借款_净额", 0) != 0:
                    bs_items.append({
                        "项目": "其他应付款-股东借款",
                        "说明": "股东借给公司的款项（负债）",
                        "本期增加": f"{profit_data['其他应付款_股东借款_收入']:,.2f}",
                        "本期减少": f"{profit_data['其他应付款_股东借款_支出']:,.2f}",
                        "净变动": f"{profit_data['其他应付款_股东借款_净额']:,.2f}",
                    })
                if profit_data.get("实收资本_净额", 0) != 0:
                    bs_items.append({
                        "项目": "实收资本",
                        "说明": "股东注资/撤资（所有者权益）",
                        "本期增加": f"{profit_data['实收资本_收入']:,.2f}",
                        "本期减少": f"{profit_data['实收资本_支出']:,.2f}",
                        "净变动": f"{profit_data['实收资本_净额']:,.2f}",
                    })
                if profit_data.get("待分类_收入", 0) > 0 or profit_data.get("待分类_支出", 0) > 0:
                    bs_items.append({
                        "项目": "待分类（需手动确认）",
                        "说明": "系统未能自动识别的交易",
                        "本期增加": f"{profit_data['待分类_收入']:,.2f}",
                        "本期减少": f"{profit_data['待分类_支出']:,.2f}",
                        "净变动": f"{profit_data['待分类_收入'] - profit_data['待分类_支出']:,.2f}",
                    })
                
                if bs_items:
                    st.subheader("📋 资产负债表项目（不计入利润）")
                    bs_df = pd.DataFrame(bs_items)
                    st.dataframe(bs_df, use_container_width=True, hide_index=True)
                    st.caption("💡 以上为资产负债表科目，不影响利润表计算。股东借款计入「其他应付款」，还款时冲减。")

                # 计算汇总（用于填入申报表）
                revenue_total = profit_data["营业收入"]
                cost_total = profit_data["营业成本"]
                profit = profit_data["利润总额"]

                expense_total = profit_data["管理费用"]
                st.subheader("📈 自动汇总结果（本期）")
                col_a, col_b, col_c, col_d = st.columns(4)
                col_a.metric("营业收入", f"{revenue_total:.2f}")
                col_b.metric("营业成本", f"{cost_total:.2f}")
                col_c.metric("管理费用", f"{expense_total:.2f}")
                col_d.metric("利润总额", f"{profit:.2f}")

                if st.button("✅ 确认并填入申报表", use_container_width=True, type="primary", key="btn_fill_quarter"):
                    st.session_state["auto_revenue"] = revenue_total
                    st.session_state["auto_cost"] = cost_total
                    st.session_state["auto_profit"] = profit
                    st.session_state["profit_data"] = profit_data  # 保存利润表数据
                    # 直接写入 widget state，绕过 st.number_input 的 value 参数不更新的 Streamlit 限制
                    st.session_state["q_revenue"] = revenue_total
                    st.session_state["q_cost"] = cost_total
                    st.session_state["q_profit"] = profit
                    st.success("✅ 已自动填入申报表，请向下滚动确认数据！")
                    st.rerun()

            except Exception as e:
                st.error(f"银行流水解析失败：{e}")
                st.caption("请确保文件包含：日期、摘要、收入金额、支出金额 等列")

    st.divider()

    # ========== 手动输入区域 ==========
    c1, c2 = st.columns(2)
    with c1:
        st.subheader("本期数（Q" + str(quarter) + "）")

        # 自动填入（如果银行流水导入了）
        rev_val = st.session_state.get("auto_revenue", 0.0)
        revenue = st.number_input("季度营业收入（元）", min_value=0.0, value=rev_val, step=1000.0, key="q_revenue")

        cost_val = st.session_state.get("auto_cost", 0.0)
        cost = st.number_input("季度营业成本（元）", min_value=0.0, value=cost_val, step=1000.0, key="q_cost")

    with c2:
        st.subheader("本期利润及企业信息")

        profit_val = st.session_state.get("auto_profit", 0.0)
        period_profit = st.number_input("季度利润总额（元）", value=profit_val, step=1000.0, key="q_profit")

        num_employees = st.number_input(
            "季度平均从业人数", min_value=1,
            value=st.session_state.get("auto_num_employees", 1), step=1, key="q_employees")
        total_assets = st.number_input(
            "季度平均资产总额（万元）", min_value=0.0,
            value=st.session_state.get("auto_total_assets", 0.0), step=10.0, key="q_assets")

    # ========== 累计数（自动计算）==========
    st.divider()
    st.subheader("📈 累计数（自动计算）")

    ytd_revenue = ytd["ytd_revenue"] + revenue
    ytd_cost = ytd["ytd_cost"] + cost
    ytd_profit = ytd["ytd_profit"] + period_profit

    col_y1, col_y2, col_y3 = st.columns(3)
    col_y1.metric("本年累计营业收入", f"{ytd_revenue:.2f} 元")
    col_y2.metric("本年累计营业成本", f"{ytd_cost:.2f} 元")
    col_y3.metric("本年累计利润总额", f"{ytd_profit:.2f} 元")

    # ========== 增值税及附加税测算 ==========
    st.divider()
    st.subheader("🧾 增值税及附加税测算")
    st.caption("依据2026年湖北省优惠政策：小规模纳税人减按1% + 六税两费减半")

    vat_col1, vat_col2 = st.columns([2, 1])
    with vat_col1:
        is_small_scale = st.radio(
            "纳税人类型",
            ["小规模纳税人（3%）", "一般纳税人（按实际税率）"],
            index=0,
            horizontal=True,
        ) == "小规模纳税人（3%）"
        vat_revenue_input = st.number_input(
            "季度含税营业收入（元，用于计算增值税）",
            min_value=0.0,
            value=float(st.session_state.get("auto_vat_revenue",
                       st.session_state.get("auto_revenue", revenue))),
            step=1000.0,
            key="vat_revenue",
            help="小规模纳税人：季度不含税收入 ≤ 30万元可免征增值税",
        )
    with vat_col2:
        st.markdown("**武汉附加税率（六税两费减半后）**")
        st.markdown("- 城建税：**3.5%**（原7%×50%）")
        st.markdown("- 教育费附加：**1.5%**（原3%×50%）")
        st.markdown("- 地方教育附加：**1%**（原2%×50%）")
        st.markdown("- 合计：**6%** × 增值税")

    # 实时预算增值税
    vat_preview = calc_vat_and_surcharge(
        revenue=vat_revenue_input,
        vat_rate=0.03,
        is_small_scale=is_small_scale,
        is_small_low_profit=True,
    )

    vc1, vc2, vc3, vc4 = st.columns(4)
    vc1.metric(
        "增值税",
        f"{vat_preview['增值税应缴']:,.2f} 元",
        delta="免税" if vat_preview['增值税应缴'] == 0 else None,
        delta_color="normal",
    )
    vc2.metric("城建税", f"{vat_preview['城建税(7%)']:,.2f} 元")
    vc3.metric("教育费附加", f"{vat_preview['教育费附加(3%)'] + vat_preview['地方教育附加(2%)']:,.2f} 元")
    vc4.metric("附加税合计", f"{vat_preview['附加税合计']:,.2f} 元")

    if vat_preview["增值税应缴"] == 0:
        st.success(f"✅ {vat_preview['增值税免税说明']}，附加税同步为零。")
    else:
        st.info(f"ℹ️ {vat_preview['增值税免税说明']}")

    # ========== 计算按钮 ==========
    st.divider()

    if st.button("🚀 计算季度预缴税额（含增值税及附加）", use_container_width=True, type="primary"):
        # ── 输入校验 ──
        q_warnings = []
        if revenue <= 0 and cost <= 0:
            q_warnings.append("⚠️ 营业收入和营业成本均为 0，请确认本季度是否有经营活动")
        if revenue > 0 and cost <= 0:
            q_warnings.append("💡 有营业收入但无营业成本，服务类企业可能正常，请核实")
        if vat_revenue_input <= 0:
            q_warnings.append("💡 增值税计算收入为 0，如本季无收入，增值税可零申报")
        if num_employees <= 0:
            q_warnings.append("⚠️ 从业人数为 0，请填写实际人数")
        if period_profit != 0 and abs(period_profit - (revenue - cost)) > revenue * 0.5:
            q_warnings.append("💡 利润总额与「收入-成本」差异较大，请确认是否已计入管理费用等期间费用")
        if q_warnings:
            with st.expander(f"🔍 数据校验提示（{len(q_warnings)} 条）", expanded=True):
                for w in q_warnings:
                    if w.startswith("⚠️"):
                        st.warning(w)
                    else:
                        st.info(w)

        # 计算本年累计已预缴税额（第12行）
        tax_paid_ytd = 0.0
        for q in range(1, quarter):
            if str(q) in prev_saved:
                tax_paid_ytd += prev_saved[str(q)].get("tax_payable", 0)

        # 增值税及附加税计算
        vat_data = calc_vat_and_surcharge(
            revenue=vat_revenue_input,
            vat_rate=0.03,
            is_small_scale=is_small_scale,
            is_small_low_profit=True,
        )
        st.session_state["vat_data"] = vat_data

        result = calc_corporate_income_tax_quarterly(
            revenue, cost, period_profit, ytd_profit,
            int(num_employees), total_assets,
            tax_paid_ytd=tax_paid_ytd,
            vat_data=vat_data,
        )
        st.session_state["corp_tax_result"] = result

        # 保存本期数据
        save_quarter_data(year, quarter, {
            "revenue": revenue,
            "cost": cost,
            "period_profit": period_profit,
            "vat_revenue": vat_revenue_input,
            "ytd_revenue": ytd_revenue,
            "ytd_cost": ytd_cost,
            "ytd_profit": ytd_profit,
            "num_employees": int(num_employees),
            "total_assets": total_assets,
            "tax_payable": result["本期应纳税额"],
            "saved_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "_source": "手动录入",
        })

        # 保存年份/季度供 Tab5 使用
        st.session_state["_tax_year"] = year
        st.session_state["_tax_quarter"] = quarter

        st.success(f"✅ 计算完成！Q{quarter} 数据已保存，下次申报 Q{quarter+1} 时会自动加载。请切换到「🏦 税款缴纳清单」查看汇总。")

    # ========== 计算结果展示（匹配 A200000 申报表格式）==========
    if "corp_tax_result" in st.session_state:
        r = st.session_state["corp_tax_result"]

        st.subheader("📋 申报表预览（A200000格式）")

        # 第一区域：收入成本利润（第1~3行）
        st.markdown("**第一部分：收入、成本、利润**")
        form_df_1 = pd.DataFrame({
            "行次": ["第1行", "第2行", "第3行"],
            "项目": ["营业收入", "营业成本", "利润总额"],
            "本期金额": [f"{r['营业收入']:,.2f}", f"{r['营业成本']:,.2f}", f"{r['利润总额']:,.2f}"],
            "累计金额": [f"{ytd_revenue:,.2f}", f"{ytd_cost:,.2f}", f"{ytd_profit:,.2f}"],
        })
        st.dataframe(form_df_1, use_container_width=True, hide_index=True)

        # 第二区域：应纳税所得额计算（第4~8行）
        st.markdown("**第二部分：应纳税所得额计算**")
        form_df_2 = pd.DataFrame({
            "行次": ["第4行", "第5行", "第6行", "第7行", "第8行"],
            "项目": ["特定业务调整", "不征税收入", "固定资产折旧调整", "弥补以前年度亏损", "实际利润额"],
            "本期金额": ["0.00", "0.00", "0.00", "0.00", f"{r['实际利润额']:,.2f}"],
            "累计金额": ["0.00", "0.00", "0.00", "0.00", f"{ytd_profit:,.2f}"],
        })
        st.dataframe(form_df_2, use_container_width=True, hide_index=True)

        # 第三区域：税款计算（第9~13行）
        st.markdown("**第三部分：税款计算**")
        form_df_3 = pd.DataFrame({
            "行次": ["第9行", "第10行", "第11行", "第12行", "第13行"],
            "项目": ["税率(25%)", "应纳所得税额", "减免所得税额", "本年累计已预缴", "本期应补(退)税额"],
            "本期金额": [
                "25%",
                f"{r['应纳税额_标准']:,.2f}",
                f"{r['减免所得税额']:,.2f}",
                f"{r['本年累计已预缴']:,.2f}",
                f"{r['本期应补(退)税额']:,.2f}",
            ],
        })
        st.dataframe(form_df_3, use_container_width=True, hide_index=True)

        # ===== 数据校验（利润表 vs 申报表）=====
        if "profit_data" in st.session_state:
            profit_data = st.session_state["profit_data"]
            validation = validate_quarterly_declaration(
                profit_data, 
                r["营业收入"], 
                r["营业成本"], 
                r["利润总额"]
            )
            
            st.subheader("📋 数据校验结果")
            all_pass = True
            for passed, msg in validation:
                if passed:
                    st.success(f"✅ {msg}")
                else:
                    st.error(f"⚠️ {msg}")
                    all_pass = False
            
            if all_pass:
                st.success("🎉 所有校验通过！利润表与申报表数据一致。")
            else:
                st.warning("⚠️ 请检查银行流水分类是否正确，或手动调整申报表数据。")

        # 关键指标卡片
        st.subheader("📊 关键指标")
        k1, k2, k3 = st.columns(3)
        k1.metric("实际利润额", f"{r['实际利润额']:,.2f} 元")
        k2.metric("企业所得税（本期应补缴）", f"{r['本期应补(退)税额']:,.2f} 元")
        k3.metric("本期税费合计", f"{r.get('本期税费合计', r['本期应补(退)税额']):,.2f} 元")

        # ===== 税费汇总表 =====
        st.subheader("💰 本期税费汇总测算")
        vat_d = st.session_state.get("vat_data", {})
        tax_summary = pd.DataFrame({
            "税种": [
                "① 企业所得税（A200000申报）",
                "② 增值税（含税收入申报）",
                "③ 城建税（增值税×7%）",
                "④ 教育费附加（增值税×3%）",
                "⑤ 地方教育附加（增值税×2%）",
                "合计应缴税费",
            ],
            "本期应缴（元）": [
                f"{r['本期应补(退)税额']:,.2f}",
                f"{vat_d.get('增值税应缴', 0.0):,.2f}",
                f"{vat_d.get('城建税(7%)', 0.0):,.2f}",
                f"{vat_d.get('教育费附加(3%)', 0.0):,.2f}",
                f"{vat_d.get('地方教育附加(2%)', 0.0):,.2f}",
                f"{r.get('本期税费合计', r['本期应补(退)税额']):,.2f}",
            ],
            "计税依据": [
                f"利润总额 {r['利润总额']:,.2f} × {'5%（小微优惠）' if r['是否小型微利企业']=='是' else '25%'}",
                vat_d.get("增值税免税说明", "-"),
                f"增值税 {vat_d.get('增值税应缴', 0.0):,.2f} × 7%",
                f"增值税 {vat_d.get('增值税应缴', 0.0):,.2f} × 3%",
                f"增值税 {vat_d.get('增值税应缴', 0.0):,.2f} × 2%",
                "① + ② + ③ + ④ + ⑤",
            ],
        })
        st.dataframe(tax_summary, use_container_width=True, hide_index=True)

        # 判断状态
        if r['利润总额'] <= 0:
            st.info("📌 本期亏损，无需缴纳企业所得税。减免所得税额和应补退税额均为0。")
        else:
            if r['是否小型微利企业'] == '是':
                st.success(f"📌 小型微利企业优惠已生效：实际税负仅5%（标准25%，减免{r['减免所得税额']:,.2f}元）")

        # AI 申报说明
        st.subheader("📄 申报说明")
        vat_d_report = st.session_state.get("vat_data", None)
        report_text = format_corporate_tax_report(r, quarter, year, vat_data=vat_d_report)
        st.text_area("申报说明", report_text, height=400, key="corp_tax_area")

        # 下载
        col_dl1, col_dl2 = st.columns(2)
        with col_dl1:
            st.download_button(
                label="📥 下载申报说明（TXT）",
                data=report_text,
                file_name=f"企业所得税预缴申报_{year}Q{quarter}.txt",
                mime="text/plain",
                use_container_width=True,
            )
        with col_dl2:
            csv_corp = pd.DataFrame([r]).to_csv(index=False, encoding="utf-8-sig")
            st.download_button(
                label="📥 下载申报底稿（CSV）",
                data=csv_corp,
                file_name=f"企业所得税预缴申报_{year}Q{quarter}.csv",
                mime="text/csv",
                use_container_width=True,
            )

    # ========== 资产负债表（会小企01表）==========
    st.divider()
    with st.expander("📋 资产负债表（会小企01表）", expanded=False):
        st.caption("依据官方 Q1 数据预填。可手动修改后用于申报信息采集。")

        bs_c1, bs_c2 = st.columns(2)

        with bs_c1:
            st.markdown("**资产**")
            cash = st.number_input("货币资金（元）", value=3466.60, step=100.0, key="bs_cash")
            st.caption(f"行1 · 流动资产合计 = {cash:,.2f}（无其他流动资产）")
            st.markdown("---")
            st.markdown(f"**资产合计** = **{cash:,.2f}** 元（行30）")

        with bs_c2:
            st.markdown("**负债和所有者权益**")
            short_loan = st.number_input("短期借款（元）", value=296.85, step=100.0, key="bs_short_loan")
            payroll_payable = st.number_input("应付职工薪酬（元）", value=11237.64, step=500.0, key="bs_payroll")
            tax_payable_bs = st.number_input("应交税费（元）", value=-150.46, step=100.0, key="bs_tax")
            other_payable = st.number_input("其他应付款（元）", value=14081.00, step=500.0, key="bs_other_pay")
            total_liab = short_loan + payroll_payable + tax_payable_bs + other_payable
            st.caption(f"行41 · 流动负债合计 = {total_liab:,.2f}（无非流动负债）")
            st.markdown("---")
            capital = st.number_input("实收资本（元）", value=45000.00, step=5000.0, key="bs_capital")
            retained = st.number_input("未分配利润（元）", value=-66998.43, step=500.0, key="bs_retained")
            equity = capital + retained
            st.caption(f"行52 · 所有者权益合计 = {equity:,.2f}")
            st.markdown("---")
            st.markdown(f"**负债和所有者权益总计** = **{total_liab + equity:,.2f}** 元（行53）")

        # 平衡校验
        assets_total = cash
        liab_equity_total = total_liab + equity
        delta = assets_total - liab_equity_total
        if abs(delta) < 1:
            st.success(f"✅ 资产负债表平衡！资产 {assets_total:,.2f} = 负债+权益 {liab_equity_total:,.2f}")
        else:
            st.error(f"⚠️ 资产负债表不平衡！资产 {assets_total:,.2f} ≠ 负债+权益 {liab_equity_total:,.2f}（差额 {delta:,.2f}）")

# ===============================================
#  Tab5：税款缴纳清单
# ===============================================

with tab4:
    st.header("🏦 税款缴纳清单 & 优惠政策适用说明")

    # 读取计算结果
    r5 = st.session_state.get("corp_tax_result", None)
    vat5 = st.session_state.get("vat_data", None)

    if r5 is None:
        st.info("💡 请先在「📊 季度申报」页面填写数据并点击「计算季度预缴税额」，完成后返回本页查看税款缴纳清单。", icon="ℹ️")
    if r5 is not None:
        # 年份/季度
        all_qdata = {}
        if os.path.exists(QUARTER_DATA_FILE):
            with open(QUARTER_DATA_FILE, "r", encoding="utf-8") as _f:
                all_qdata = json.load(_f)

        latest_year = max(all_qdata.keys(), default=str(datetime.now().year))
        latest_qdata_key = str(st.session_state.get("_tax_year", int(latest_year)))
        disp_year = st.session_state.get("_tax_year", int(latest_year))
        disp_quarter = st.session_state.get("_tax_quarter", datetime.now().month // 3 or 1)

        # ===== 工资个税数据 =====
        payroll_results = st.session_state.get("results", [])
        total_personal_tax = sum(float(e.get("应纳税额", 0)) for e in payroll_results)
        total_company_social = sum(float(e.get("公司社保", 0)) for e in payroll_results)
        total_personal_social = sum(float(e.get("个人社保", 0)) for e in payroll_results)
        employee_count = len(payroll_results)

        # ===== 增值税及附加 =====
        vat_amount = vat5.get("增值税应缴", 0.0) if vat5 else 0.0
        urban_tax = vat5.get("城建税(7%)", 0.0) if vat5 else 0.0
        edu_surcharge = vat5.get("教育费附加(3%)", 0.0) if vat5 else 0.0
        local_edu = vat5.get("地方教育附加(2%)", 0.0) if vat5 else 0.0
        surcharge_total = vat5.get("附加税合计", 0.0) if vat5 else 0.0
        vat_note = vat5.get("增值税免税说明", "-") if vat5 else "未计算"
        six_two_relief = vat5.get("六税两费减免金额", 0.0) if vat5 else 0.0
        vat_policy = vat5.get("增值税优惠依据", "-") if vat5 else "-"
        surcharge_policy = vat5.get("附加税优惠说明", "") if vat5 else ""

        # ===== 企业所得税 =====
        corp_tax = r5.get("本期应补(退)税额", 0.0) if r5 else 0.0
        corp_relief = r5.get("减免所得税额", 0.0) if r5 else 0.0
        is_small = (r5.get("是否小型微利企业", "否") == "是") if r5 else True

        # ===== 印花税 =====
        # 资金账簿：注册资本到位 / 增资
        stamp_reg_capital = st.session_state.get("stamp_reg_capital", 0.0)
        stamp_capital_increase = st.session_state.get("stamp_capital_increase", 0.0)
        # 购销合同：按当季收入估算（可视收入为含税购销额）
        stamp_purchase = vat5.get("季度含税收入", 0.0) if vat5 else 0.0
        stamp_data = calc_stamp_duty(
            registered_capital=stamp_reg_capital,
            capital_increase=stamp_capital_increase,
            purchase_amount=stamp_purchase,
            is_small_low_profit=is_small,
        )
        stamp_total = stamp_data["印花税合计（应缴）"]
        stamp_nominal = stamp_data["印花税合计（名义）"]
        stamp_relief = stamp_data["六税两费减免"]

        # ===== 合计（含印花税）=====
        total_tax = round(total_personal_tax + vat_amount + surcharge_total + corp_tax + stamp_total, 2)

        # ===== 1. 优惠政策适用清单（先展示） =====
        st.subheader(f"📋 {disp_year}年第{disp_quarter}季度 — 优惠税率适用说明")
        st.caption(f"政策依据：湖北省《关于加力助企解难推动中小企业稳健发展的若干措施》（有效期至2027.12.31）")

        # 获取政策汇总
        revenue_used = st.session_state.get("vat_revenue", r5.get("营业收入", 0.0) if r5 else 0.0)
        emp_used = r5.get("从业人数", 1) if r5 else 1
        asset_used = r5.get("资产总额_万元", 0.0) if r5 else 0.0

        policy_summary = get_tax_policy_summary(
            is_small_scale=(vat5.get("是否小规模纳税人", "是") == "是") if vat5 else True,
            is_small_low_profit=is_small,
            quarter_revenue=revenue_used,
            num_employees=emp_used,
            total_assets=asset_used,
            quarter=disp_quarter,
        )

        for i, p in enumerate(policy_summary["policies"], 1):
            with st.expander(f"🎯 {p['税种']} — {p['优惠名称']}", expanded=(i == 1)):
                cols = st.columns([2, 1])
                with cols[0]:
                    st.markdown(f"""
    - **优惠内容**：{p['优惠内容']}
    - **政策依据**：{p['政策依据']}
    - **适用条件**：{p['适用条件']}
                    """)
                with cols[1]:
                    st.metric("优惠力度", p['优惠力度'])

        st.success(f"✅ {policy_summary['tip']}")
        st.info(f"📅 {policy_summary['valid_until']}")

        st.divider()

        # ===== 2. 税款缴纳清单（含优惠标注） =====
        st.subheader("📋 税款明细清单（含优惠税率标注）")

        # 确定增值税的标称值和优惠说明
        vat_display_rate = "0%（免税）" if vat_amount == 0 else "1%（优惠后，原3%）"
        if not (vat5 and vat5.get("是否小规模纳税人") == "是"):
            vat_display_rate = f"{vat5.get('增值税名义税率', 0.03)*100:.0f}%（一般纳税人）" if vat5 else "-"

        rows = [
            {
                "序号": "1",
                "税款类型": "🏢 增值税",
                "标称税率": "3%（小规模）",
                "优惠后税率": vat_display_rate,
                "适用优惠": "小规模减按1% + 季≤30万免税" if vat_amount == 0 else "小规模减按1%征收",
                "政策依据": "财税〔2023〕19号",
                "本期应缴（元）": f"{vat_amount:,.2f}",
                "状态": "免税 ✅" if vat_amount == 0 else "待缴 ⏳",
            },
            {
                "序号": "2",
                "税款类型": "🏙️ 城建税",
                "标称税率": "7%（市区）",
                "优惠后税率": "3.5%（减半）" if (vat5 and vat5.get("是否享受六税两费减半") == "是") else "7%",
                "适用优惠": "「六税两费」减半征收" if (vat5 and vat5.get("是否享受六税两费减半") == "是") else "无",
                "政策依据": "财税〔2022〕10号",
                "本期应缴（元）": f"{urban_tax:,.2f}",
                "状态": "免税 ✅" if urban_tax == 0 else "待缴 ⏳",
            },
            {
                "序号": "3",
                "税款类型": "📚 教育费附加",
                "标称税率": "3%",
                "优惠后税率": "1.5%（减半）" if (vat5 and vat5.get("是否享受六税两费减半") == "是") else "3%",
                "适用优惠": "「六税两费」减半征收",
                "政策依据": "财税〔2022〕10号",
                "本期应缴（元）": f"{edu_surcharge:,.2f}",
                "状态": "免税 ✅" if edu_surcharge == 0 else "待缴 ⏳",
            },
            {
                "序号": "4",
                "税款类型": "🎓 地方教育附加",
                "标称税率": "2%",
                "优惠后税率": "1%（减半）" if (vat5 and vat5.get("是否享受六税两费减半") == "是") else "2%",
                "适用优惠": "「六税两费」减半征收",
                "政策依据": "财税〔2022〕10号",
                "本期应缴（元）": f"{local_edu:,.2f}",
                "状态": "免税 ✅" if local_edu == 0 else "待缴 ⏳",
            },
            {
                "序号": "5",
                "税款类型": "💼 企业所得税（季预缴）",
                "标称税率": "25%",
                "优惠后税率": "5%（小微优惠）" if is_small else "25%",
                "适用优惠": "小型微利企业：减按25%计税×20%税率=5%" if is_small else "不适用小微优惠",
                "政策依据": "财税〔2023〕12号",
                "本期应缴（元）": f"{corp_tax:,.2f}",
                "状态": "无需缴纳 ✅" if corp_tax == 0 else "待缴 ⏳",
            },
            {
                "序号": "6",
                "税款类型": "👤 个人所得税（代扣代缴）",
                "标称税率": "3%-45%累进",
                "优惠后税率": "3%-45%（起征点5000元/月）",
                "适用优惠": f"基本减除5000元+专项附加扣除（{employee_count}名员工）",
                "政策依据": "个人所得税法",
                "本期应缴（元）": f"{total_personal_tax:,.2f}",
                "状态": "无个税 ✅" if total_personal_tax == 0 else "待缴 ⏳",
            },
        ]

        # 印花税明细行
        for si in stamp_data.get("明细", []):
            rows.append({
                "序号": f"7-{stamp_data['明细'].index(si)+1}" if len(stamp_data.get("明细", [])) > 1 else "7",
                "税款类型": f"📜 印花税-{si['税目']}",
                "标称税率": si["名义税率"],
                "优惠后税率": si["优惠后税率"],
                "适用优惠": "「六税两费」减半征收" if stamp_data["是否六税两费减半"] == "是" else "标准税率",
                "政策依据": "印花税法 + 财税〔2022〕10号",
                "本期应缴（元）": f"{si['应纳税额（元）']:,.2f}",
                "状态": "免税 ✅" if si['应纳税额（元）'] == 0 else "待缴 ⏳",
            })

        if stamp_data["税目数量"] == 0:
            rows.append({
                "序号": "7",
                "税款类型": "📜 印花税",
                "标称税率": "见各税目",
                "优惠后税率": "六税两费减半",
                "适用优惠": "小型微利 → 各税目减半",
                "政策依据": "印花税法（2022.7.1）",
                "本期应缴（元）": "0.00",
                "状态": "本季无需缴纳 ✅",
            })

        df_tax_list = pd.DataFrame(rows)
        st.dataframe(
            df_tax_list.style.apply(
                lambda row: ["background-color: #e8f5e9" if "免税" in str(row["状态"]) or "无需缴纳" in str(row["状态"]) else
                             ("background-color: #fff9c4" if "待缴" in str(row["状态"]) else "")
                             for _ in row],
                axis=1
            ),
            use_container_width=True,
            hide_index=True,
        )

        st.caption("🟢 绿色行 = 享受优惠后无需缴纳 | 🟡 黄色行 = 需按期缴纳")

        # ===== 3. 六税两费减半明细 =====
        if six_two_relief > 0 or stamp_relief > 0:
            st.divider()
            st.subheader("🎁 「六税两费」减半征收 — 本季减免明细")
            col_count = 3 + (1 if stamp_relief > 0 else 0)
            relief_cols = st.columns(col_count)
            relief_idx = 0
            relief_cols[relief_idx].metric("城建税减免", f"{vat5.get('城建税名义', 0) - urban_tax:,.2f} 元"); relief_idx += 1
            relief_cols[relief_idx].metric("教育费附加减免", f"{vat5.get('教育费附加名义', 0) - edu_surcharge:,.2f} 元"); relief_idx += 1
            relief_cols[relief_idx].metric("地方教育附加减免", f"{vat5.get('地方教育附加名义', 0) - local_edu:,.2f} 元"); relief_idx += 1
            if stamp_relief > 0:
                relief_cols[relief_idx].metric("印花税减免", f"{stamp_relief:,.2f} 元")
            st.success(f"💰 本季「六税两费」合计减免：**{six_two_relief + stamp_relief:,.2f} 元**")

        # ===== 4. 汇总指标卡 =====
        st.divider()
        st.subheader("💰 本期税款汇总（优惠后）")

        mc1, mc2, mc3, mc4, mc5 = st.columns(5)
        mc1.metric(
            "增值税及附加",
            f"{round(vat_amount + surcharge_total, 2):,.2f} 元",
            delta="全部免税" if vat_amount == 0 else None,
        )
        mc2.metric(
            "企业所得税",
            f"{corp_tax:,.2f} 元",
            delta=f"减免 {corp_relief:,.0f} 元" if corp_relief > 0 else ("亏损无需缴纳" if corp_tax == 0 else None),
        )
        mc3.metric(
            "个人所得税",
            f"{total_personal_tax:,.2f} 元",
            delta=f"{employee_count}名员工" if employee_count > 0 else None,
        )
        mc4.metric(
            "📜 印花税",
            f"{stamp_total:,.2f} 元",
            delta=f"减免 {stamp_relief:,.0f} 元" if stamp_relief > 0 else "本季无需缴纳",
        )
        mc5.metric(
            "🔴 本期税款合计",
            f"{total_tax:,.2f} 元",
            delta=f"六税两费减免 {six_two_relief + stamp_relief:,.0f}" if (six_two_relief + stamp_relief) > 0 else None,
        )

        # ===== 5. 社保提醒 =====
        if total_company_social > 0 or total_personal_social > 0:
            st.divider()
            st.subheader("🛡️ 社保缴费提醒（非税款，单独缴纳）")
            s1, s2, s3 = st.columns(3)
            s1.metric("公司承担社保", f"{total_company_social:,.2f} 元")
            s2.metric("个人承担社保", f"{total_personal_social:,.2f} 元")
            s3.metric("社保合计", f"{total_company_social + total_personal_social:,.2f} 元")
            st.info("💡 社保缴费请登录**武汉市社会保险网上服务平台**或通过银行代扣完成，截止日期一般为当月25日。")

        # ===== 6. 残保金 =====
        if employee_count <= 30:
            st.divider()
            st.subheader("♿ 残疾人就业保障金")
            st.success(f"✅ 在职职工 {employee_count} 人 ≤ 30人 → **免征残疾人就业保障金**（发改价格规〔2019〕2015号）")

        # ===== 7. 缴款期限 =====
        st.divider()
        st.subheader("⏰ 缴款期限提醒")

        q_end_month = {1: 3, 2: 6, 3: 9, 4: 12}
        deadline_month = q_end_month.get(disp_quarter, 3) + 1
        if deadline_month > 12:
            deadline_month = 1
            deadline_year = disp_year + 1
        else:
            deadline_year = disp_year

        st.markdown(f"""
    | 税种 | 优惠政策 | 实际税率 | 申报截止日期 | 申报平台 |
    |------|---------|---------|------------|---------|
    | 增值税 | 小规模减按1% | {vat_display_rate} | **{deadline_year}年{deadline_month}月15日** | 湖北省电子税务局 |
    | 城建税 | 六税两费减半 | 3.5% | 同上 | 随增值税一并申报 |
    | 教育费附加+地方教育附加 | 六税两费减半 | 1.5%+1% | 同上 | 随增值税一并申报 |
    | 企业所得税 | 小型微利5% | {'5%' if is_small else '25%'} | **{deadline_year}年{deadline_month}月15日** | 电子税务局 → A200000 |
    | 个人所得税 | 起征点5000 | 3%-45% | **次月15日** | 自然人税收管理系统 |
    | 印花税 | 六税两费减半 | 各税目减半 | 按次/按期汇总 | 湖北省电子税务局 |
    | 残保金 | ≤30人免征 | 0% | 年度申报 | 残联/税务部门 |
        """)

        st.markdown("""
    > **📌 操作步骤：**
    > 1. 登录 [湖北省电子税务局](https://etax.hubei.chinatax.gov.cn/) 完成增值税及企业所得税申报
    > 2. 申报完成后，通过网银或第三方支付完成税款划缴
    > 3. 截图留存申报成功页面，归入税务档案
    > 4. 个税通过**自然人税收管理系统（扣缴客户端）**申报并缴纳
    > 5. 优惠政策**无需额外申请**，系统自动识别减免（湖北省「免申即享」）
        """)

        # ===== 8. 下载 =====
        st.divider()
        dl1, dl2 = st.columns(2)
        with dl1:
            csv_tax = df_tax_list.to_csv(index=False, encoding="utf-8-sig")
            st.download_button(
                label="📥 下载税款缴纳清单（CSV）",
                data=csv_tax,
                file_name=f"税款缴纳清单_{disp_year}Q{disp_quarter}.csv",
                mime="text/csv",
                use_container_width=True,
            )
        with dl2:
            summary_lines = [
                f"湖北启贤托儿所有限公司",
                f"{disp_year}年第{disp_quarter}季度 税款缴纳清单（含优惠政策）",
                f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M')}",
                "=" * 50,
                f"▶ 适用优惠政策：",
                f"  1. 小规模纳税人季度免税/减按1%（季≤30万免，超则1%）",
                f"  2. 六税两费减半征收（城建/教育/地方教育/印花税）",
                f"  3. 小型微利企业所得税5%（年利润≤300万）",
                f"  4. 残保金免征（员工≤30人）",
                "=" * 50,
                f"▶ 本期实际应缴：",
                f"  增值税：{vat_amount:,.2f} 元 ({vat_display_rate})",
                f"  城建税：{urban_tax:,.2f} 元 (3.5%)",
                f"  教育费附加：{edu_surcharge:,.2f} 元 (1.5%)",
                f"  地方教育附加：{local_edu:,.2f} 元 (1%)",
                f"  企业所得税：{corp_tax:,.2f} 元 ({'5%' if is_small else '25%'})",
                f"  个人所得税：{total_personal_tax:,.2f} 元",
                f"  印花税：{stamp_total:,.2f} 元（含资金账簿/购销合同等，减半后）",
                "-" * 30,
                f"  本期税款合计：{total_tax:,.2f} 元",
                f"  六税两费减免（含印花税）：{six_two_relief + stamp_relief:,.2f} 元",
                f"  企业所得税减免：{corp_relief:,.2f} 元",
                "=" * 50,
                f"申报截止：{deadline_year}年{deadline_month}月15日（增值税、企业所得税）",
                f"个税截止：次月15日",
                f"社保截止：当月25日",
                f"印花税：按次或按期汇总申报",
                f"政策有效期至：2027年12月31日",
            ]
            summary_txt = "\n".join(summary_lines)
            st.download_button(
                label="📥 下载缴纳通知书（TXT）",
                data=summary_txt,
                file_name=f"税款缴纳通知书_{disp_year}Q{disp_quarter}.txt",
                mime="text/plain",
                use_container_width=True,
            )

    # ===============================================
    #  Tab6：申报指南（四合一）
    # ===============================================

with tab7:
    st.header("📖 湖北启贤托儿所有限公司 — 税务申报操作指南")
    st.caption("基于小规模纳税人、小型微利企业场景 · 湖北省2026年度适用")

    guide_choice = st.radio(
        "请选择要查阅的申报说明：",
        [
            "一、个人所得税申报（代扣代缴）",
            "二、企业所得税季报 & 年报",
            "三、利润表编制说明",
            "四、增值税及附加税等税目申报",
        ],
        horizontal=True,
    )

    st.divider()

    # ==========================================
    #  一、个人所得税申报
    # ==========================================
    if guide_choice.startswith("一、"):
        st.subheader("一、个人所得税申报（代扣代缴）")

        with st.expander("1.1 申报主体与义务", expanded=True):
            st.markdown("""
**申报主体**：湖北启贤托儿所有限公司（作为「扣缴义务人」）

**法律依据**：《中华人民共和国个人所得税法》第九条 —— 个人所得税以所得人为纳税人，以支付所得的单位或者个人为扣缴义务人。

**申报内容**：
- 工资、薪金所得（员工每月工资）
- 劳务报酬所得（如有外部人员）
- 公司需在发放工资时依法代扣代缴个税，按月向税务机关申报

**关键概念**：
| 术语 | 说明 |
|------|------|
| 扣缴义务人 | 公司（代扣代缴） |
| 纳税人 | 员工个人 |
| 累计预扣法 | 按月累计计算，全年汇算 |
| 专项附加扣除 | 子女教育/婴幼儿照护/赡养老人/住房贷款/住房租金/继续教育/大病医疗 |
            """)

        with st.expander("1.2 计算公式（累计预扣法）", expanded=False):
            st.markdown("""
**核心公式：**

本期应预扣预缴税额 = （累计预扣预缴应纳税所得额 × 预扣率 - 速算扣除数） - 累计减免税额 - 累计已预扣预缴税额

其中：
- 累计预扣预缴应纳税所得额 = 累计收入 - 累计免税收入 - 累计减除费用 - 累计专项扣除 - 累计专项附加扣除 - 累计依法确定的其他扣除

**月度计算公式（简化）：**
```
应税收入 = 税前工资 - 个人社保 - 基本减除费用(5000元) - 专项附加扣除
应纳税额 = 应税收入 × 适用税率 - 速算扣除数
实发工资 = 税前工资 - 个人社保 - 应纳税额
```

**7级超额累进税率表（2024-2026年）：**

| 级数 | 累计应纳税所得额 | 税率 | 速算扣除数 |
|------|-----------------|------|-----------|
| 1 | ≤ 3,000 | 3% | 0 |
| 2 | 3,001 ~ 12,000 | 10% | 210 |
| 3 | 12,001 ~ 25,000 | 20% | 1,410 |
| 4 | 25,001 ~ 35,000 | 25% | 2,660 |
| 5 | 35,001 ~ 55,000 | 30% | 4,410 |
| 6 | 55,001 ~ 80,000 | 35% | 7,160 |
| 7 | > 80,000 | 45% | 15,160 |

**示例计算**（员工A，税前10522元，个人社保522元，专项附加扣除5000元）：
```
应税收入 = 10522 - 522 - 5000 - 5000 = 0 元
应纳税额 = 0 元（无需缴纳个税，需进行零申报）
实发工资 = 10522 - 522 = 10,000 元
```
            """)

        with st.expander("1.3 专项附加扣除明细", expanded=False):
            st.markdown("""
**2026年武汉地区适用标准：**

| 扣除项目 | 每月标准 | 说明 |
|----------|---------|------|
| 子女教育 | 2,000元/人 | 3岁至博士均可，父母各扣50%或一方全扣 |
| 婴幼儿照护 | 2,000元/人 | 0~3岁婴幼儿 |
| 赡养老人 | 最高3,000元 | 60岁以上父母，独生子女3000元，非独生子女分摊 |
| 住房贷款利息 | 1,000元 | 首套房贷款 |
| 住房租金 | 800~1,500元 | 武汉属省会城市，按1,500元/月 |
| 继续教育 | 400元/月 | 学历继续教育期间 |
| 大病医疗 | 据实扣除 | 年度汇算清缴时扣除 |

> **重要提醒**：员工需在「个人所得税APP」中自行填报专项附加扣除信息，公司端（扣缴端）自动同步。如有变动请提醒员工及时更新。
            """)

        with st.expander("1.4 申报操作流程", expanded=False):
            st.markdown("""
**申报平台**：自然人电子税务局（扣缴端） → https://etax.chinatax.gov.cn/

**操作步骤（每月一次）：**

```
第1步：人员信息采集
  ├─ 登录扣缴端 → 「人员信息采集」 → 添加/导入员工信息
  ├─ 必填：姓名、身份证号、任职受雇从业类型（雇员）
  └─ 报送 → 获取反馈（验证通过后可用）

第2步：专项附加扣除信息采集
  ├─ 「专项附加扣除信息采集」 → 「下载更新」
  ├─ 系统自动从税务局同步员工在APP中填报的信息
  └─ 核对是否有新增/变更的员工扣除信息

第3步：综合所得申报
  ├─ 「综合所得申报」 → 「收入及减除填写」 → 「正常工资薪金所得」
  ├─ 导入或手动填写：本期收入、个人社保、专项附加扣除
  ├─ 「税款计算」 → 系统自动计算每位员工应缴个税
  └─ 「申报表报送」 → 核对无误后发送申报

第4步：税款缴纳
  ├─ 「税款缴纳」 → 选择三方协议扣款或银行端查询缴税
  └─ 扣款成功后截图留存凭证
```

**零申报操作**（如所有员工应税收入为0）：
- 同样走完整流程，税款计算结果为0元
- 仍需点击「申报表报送」完成零申报
- **切勿遗漏零申报**，否则会产生逾期未申报记录
            """)

        with st.expander("1.5 申报周期与截止日期", expanded=False):
            st.markdown("""
| 事项 | 周期 | 截止时间 | 说明 |
|------|------|---------|------|
| 月度代扣代缴 | 每月 | 次月15日 | 例：5月工资 → 6月15日前申报 |
| 年度汇算清缴 | 每年 | 次年3月1日~6月30日 | 员工自行在个税APP操作，公司无需代办 |
| 扣缴端人员信息更新 | 有变动时 | 当月申报前 | 新入职/离职员工需及时更新 |

**逾期后果**：
- 逾期申报：按日加收滞纳税款万分之五的滞纳金
- 连续逾期：可能被列入重点监管名单
            """)

        with st.expander("1.6 常见问题", expanded=False):
            st.markdown("""
**Q1：员工工资很低，还需要申报吗？**
A：需要。即使应税收入为0、无需缴税，也必须进行零申报。

**Q2：专项附加扣除谁填？**
A：员工在个人所得税APP中自行填报，公司在扣缴端「下载更新」即可同步。

**Q3：社保基数变了怎么办？**
A：每年社保基数调整后（通常7月），在扣缴端修改员工「个人社保」金额即可。

**Q4：全年一次性奖金（年终奖）怎么处理？**
A：可选择单独计税或并入综合所得。建议测算两种方式后选择税额更低的方案。

**Q5：离职员工怎么处理？**
A：在「人员信息采集」中将状态改为「非正常」，填写离职日期。
            """)

    # ==========================================
    #  二、企业所得税季报 & 年报
    # ==========================================
    elif guide_choice.startswith("二、"):
        st.subheader("二、企业所得税季报 & 年报")

        with st.expander("2.1 申报类型与周期总览", expanded=True):
            st.markdown("""
**企业所得税申报分为两类：**

| 申报类型 | 周期 | 表单 | 截止时间 | 说明 |
|----------|------|------|---------|------|
| **季度预缴** | 每季度 | A200000表 | 季度终了后15日内 | 按季度利润预缴 |
| **年度汇算清缴** | 每年 | A类年度申报表 | 次年5月31日前 | 全年汇总多退少补 |

**2026年申报日历：**

| 期间 | 申报内容 | 截止日期 |
|------|---------|---------|
| Q1（1-3月） | 季度预缴 | **2026年4月15日** |
| Q2（4-6月） | 季度预缴 | **2026年7月15日** |
| Q3（7-9月） | 季度预缴 | **2026年10月15日** |
| Q4（10-12月） | 季度预缴 | **2027年1月15日** |
| 2026全年 | 年度汇算清缴 | **2027年5月31日** |

> ⚠️ **重要**：季度预缴不能少报或漏报，否则年度汇算时会被追缴并加收滞纳金。

**此外还有「工商年报」（非税务）：**
- 申报平台：国家企业信用信息公示系统
- 申报时间：每年1月1日 ~ 6月30日
- 内容：企业基本信息、股东出资、资产状况、社保信息等
            """)

        with st.expander("2.2 季度预缴申报（A200000表）", expanded=False):
            st.markdown("""
**申报平台**：湖北省电子税务局 → https://etax.hubei.chinatax.gov.cn/

**A200000表核心行次：**

```
第1行  营业收入          = 主营业务收入 + 其他业务收入
第2行  营业成本          = 主营业务成本 + 其他业务成本
第3行  利润总额          = 营业收入 - 营业成本 - 税金及附加 - 管理费用 - 财务费用 + 投资收益 + 营业外收入 - 营业外支出
第4行  特定业务调整      = 0（一般企业无需填写）
第5行  不征税收入        = 0（如有政府补贴等需注明）
第6行  固定资产折旧调整  = 0（默认无差异）
第7行  弥补以前年度亏损  = 以前年度未弥补亏损额
第8行  实际利润额        = 第3行 + 第4行 - 第5行 - 第6行 - 第7行
第9行  税率              = 25%（固定）
第10行 应纳所得税额      = 第8行 × 25%
第11行 减免所得税额      = 小型微利企业优惠减免
第12行 本年累计已预缴    = 累计之前季度已缴税额
第13行 本期应补(退)税额  = 本期应纳税额 - 第12行
```

**小型微利企业优惠（Q1~Q4自动适用）：**
- 条件：年利润≤300万 + 员工≤300人 + 资产≤5000万
- 实际税率：**5%**（减按25%计入应纳税所得额 × 20%税率）
- 湖北启贤托儿所符合条件 ✅

**季度预缴与利润表的关系：**
```
利润表中的「利润总额(第12行)」 ≈ 申报表「第3行 利润总额」

差异说明：
- 如果利润表按小企业会计准则正常编制，可直接填入申报表
- 如有纳税调整项（如不可扣除费用），在年度汇算时调整，季度预缴按账面利润
```
            """)

        with st.expander("2.3 年度汇算清缴", expanded=False):
            st.markdown("""
**申报平台**：湖北省电子税务局 → 「企业所得税年度申报」

**A类年度申报表结构：**

| 部分 | 内容 | 关键表 |
|------|------|--------|
| 主表 | A100000 年度纳税申报表 | 汇总所有数据 |
| 收入类 | A101010 一般企业收入明细表 | 营业收入/营业外收入 |
| 成本类 | A102010 一般企业成本支出明细表 | 营业成本/期间费用 |
| 期间费用 | A104000 期间费用明细表 | 管理/销售/财务费用 |
| 纳税调整 | A105000 纳税调整项目明细表 | 调增/调减项目 |
| 优惠类 | A107040 减免所得税优惠明细表 | 小型微利企业等 |
| 弥补亏损 | A106000 弥补亏损明细表 | 以前年度亏损 |
| 职工薪酬 | A105050 职工薪酬支出及纳税调整明细表 | 工资/福利/社保 |

**汇算清缴核心逻辑：**
```
会计利润总额（利润表）
  + 纳税调增项（超标费用、不得扣除支出等）
  - 纳税调减项（免税收入、加计扣除等）
  - 弥补以前年度亏损
  = 应纳税所得额
  × 适用税率（小型微利5%）
  = 全年应纳税额
  - 已预缴税额（Q1~Q4合计）
  = 应补(退)税额
```

**常见纳税调整项：**

| 项目 | 调整方向 | 说明 |
|------|---------|------|
| 业务招待费 | 调增 | 按发生额60%扣除，最高不超过营业收入5‰ |
| 广告宣传费 | 可能调增 | 不超过营业收入15%，超过部分可结转 |
| 罚款/滞纳金 | 调增 | 不得税前扣除 |
| 未取得发票的费用 | 调增 | 需取得合规发票 |
| 残疾人员工资 | 调减 | 可加计100%扣除 |
            """)

        with st.expander("2.4 小型微利企业判定与优惠", expanded=False):
            st.markdown("""
**小型微利企业「335标准」：**

| 条件 | 标准 | 湖北启贤托儿所 | 判断 |
|------|------|---------------|------|
| 年应纳税所得额 | ≤ 300万元 | 取决于当年利润 | ✅/⚠️ |
| 从业人数 | ≤ 300人 | 通常1-5人 | ✅ |
| 资产总额 | ≤ 5,000万元 | 通常较小 | ✅ |

**优惠政策（2024-2027年）：**

```
实际税负 = 25%（减按计税比例） × 20%（优惠税率） = 5%

🔴 举例：
  全年利润 100万元
  应纳税所得额 = 100万（假设无调整项）
  应纳税额（标准）= 100万 × 25% = 25万
  应纳税额（优惠）= 100万 × 5% = 5万
  减免税额 = 25万 - 5万 = 20万 ✅
```

**注意事项：**
- 优惠政策无需额外申请，系统自动判定
- 湖北省已实现「免申即享」
- 如年度利润超过300万，当年不再享受小微优惠，需按25%缴纳
            """)

        with st.expander("2.5 申报操作流程", expanded=False):
            st.markdown("""
**季度预缴操作步骤（每季度一次）：**

```
第1步：登录 → 湖北省电子税务局 → 企业登录（税号+密码或CA证书）
第2步：进入 → 「我要办税」 → 「税费申报及缴纳」
第3步：选择 → 「企业所得税申报」 → 「居民企业（查账征收）所得税月（季）度申报」
第4步：填报 → 按A200000表格式填写收入/成本/利润
第5步：系统自动判定小型微利企业并计算减免税额
第6步：预览 → 核对与利润表是否一致
第7步：申报 → 点击「申报」→ 确认提交
第8步：缴款 → 如有应缴税款，通过三方协议或网银缴款
第9步：留存 → 截图申报成功页面，保存PDF回执
```

**年度汇算清缴操作步骤（每年一次，5月31日前）：**

```
第1步：登录 → 湖北省电子税务局
第2步：进入 → 「企业所得税年度申报」
第3步：先填附表（A101010~A107040），再填主表A100000
第4步：重点核对：纳税调整明细表A105000
第5步：系统自动计算应补退税额
第6步：申报并缴款（如有补税）
第7步：留存全套申报表PDF
```
            """)

    # ==========================================
    #  三、利润表编制说明
    # ==========================================
    elif guide_choice.startswith("三、"):
        st.subheader("三、利润表编制说明（小企业会计准则）")

        with st.expander("3.1 会计准则与适用范围", expanded=True):
            st.markdown("""
**湖北启贤托儿所有限公司适用：《小企业会计准则》**

| 准则类型 | 适用对象 | 湖北启贤托儿所 |
|----------|---------|---------------|
| 企业会计准则 | 大中型企业、上市公司 | ❌ |
| 小企业会计准则 | 小微型企业 | ✅ |
| 个体工商户会计制度 | 个体户 | ❌ |

**小企业会计准则要点：**
- 简化核算，无需计提减值准备
- 资产按历史成本计量
- 所得税费用采用应付税款法（不确认递延所得税）
- 利润表结构比企业会计准则更简洁
            """)

        with st.expander("3.2 利润表项目逐项解释", expanded=False):
            st.markdown("""
**小企业会计准则利润表（标准格式）：**

| 行次 | 项目 | 含义 | 数据来源 |
|------|------|------|---------|
| 第1行 | **一、营业收入** | 主营业务收入 + 其他业务收入 | 银行流水+发票 |
| 第2行 | 减：营业成本 | 主营业务成本 + 其他业务成本 | 采购/进货/材料成本 |
| 第3行 | 税金及附加 | 城建税+教育费附加+地方教育附加+印花税+房产税等 | 税务申报表 |
| 第4行 | 销售费用 | 销售相关费用（广告、运输等） | 银行流水分类 |
| 第5行 | 管理费用 | 工资社保+办公费+差旅+招待等 | 银行流水+工资表 |
| 第6行 | 财务费用 | 利息收支+银行手续费 | 银行流水 |
| 第7行 | 资产减值损失 | 小企业会计准则下通常为0 | - |
| 第8行 | 加：投资收益 | 对外投资取得的分红/收益 | 银行流水 |
| 第9行 | **二、营业利润** | = 第1行 - 第2行 - 第3行 - 第4行 - 第5行 - 第6行 - 第7行 + 第8行 | 计算 |
| 第10行 | 加：营业外收入 | 政府补助/罚款收入/盘盈等 | 银行流水 |
| 第11行 | 减：营业外支出 | 罚款支出/捐赠/资产损失 | 银行流水 |
| 第12行 | **三、利润总额** | = 第9行 + 第10行 - 第11行 | 计算 |
| 第13行 | 减：所得税费用 | 利润总额 × 5%（小型微利） | 计算 |
| 第14行 | **四、净利润** | = 第12行 - 第13行 | 最终成果 |

**⚠️ 关键校验公式：**
```
利润总额(第12行) = 企业所得税预缴申报表A200000 第3行
营业收入(第1行) = A200000 第1行
营业成本(第2行) = A200000 第2行
```
            """)

        with st.expander("3.3 银行流水 → 利润表 映射表", expanded=False):
            st.markdown("""
**从银行流水到利润表的自动分类逻辑：**

| 银行流水关键词 | → 利润表项目 | → 申报表行次 |
|---------------|-------------|------------|
| 货款/销售收入/服务费/咨询费/收款 | **营业收入** | 第1行 |
| 采购/进货/材料成本 | **营业成本** | 第2行 |
| 税金/城建税/教育费附加/印花税 | **税金及附加** | 第3行 |
| 运输费/广告费/推广费 | **销售费用** | 第4行 |
| 工资/社保/办公/差旅/招待/房租/水电 | **管理费用** | 第5行 |
| 利息/银行手续费 | **财务费用** | 第6行 |
| 投资收益/分红/股息 | **投资收益** | 第8行 |
| 政府补助/罚款收入 | **营业外收入** | 第10行 |
| 罚款支出/捐赠 | **营业外支出** | 第11行 |

**注意事项：**
- 银行流水需覆盖全部收支，如有现金交易需补充登记
- 工资社保数据需从工资表导入，银行流水仅反映实发金额
- 税金及附加需从税务申报表取数，银行流水中的税款支出仅供参考
            """)

        with st.expander("3.4 利润表编制步骤", expanded=False):
            st.markdown("""
**月度/季度利润表编制流程：**

```
第1步：收集原始凭证
  ├─ 银行流水（所有对公账户）
  ├─ 工资发放表（含社保明细）
  ├─ 发票台账（收入/支出）
  └─ 税务缴纳记录

第2步：分类汇总
  ├─ 使用本系统的「银行流水自动分类」功能
  ├─ 或手动按利润表项目归类
  └─ 确保每一笔交易都有正确的利润表归属

第3步：填表计算
  ├─ 各项目汇总金额填入对应行次
  ├─ 计算营业利润（第9行）= 收入 - 成本 - 税费 - 费用 + 投资收益
  ├─ 计算利润总额（第12行）= 营业利润 + 营业外收入 - 营业外支出
  └─ 计算所得税（第13行）= 利润总额 × 5%

第4步：交叉校验
  ├─ 利润表第12行（利润总额） = 申报表A200000第3行
  ├─ 利润表第1行（营业收入） = 申报表A200000第1行
  └─ 利润表第2行（营业成本） = 申报表A200000第2行

第5步：归档留存
  ├─ 利润表原件（Excel/PDF）
  ├─ 银行流水完整版
  ├─ 发票清单
  └─ 申报回执
```
            """)

        with st.expander("3.5 常见错误与纠正", expanded=False):
            st.markdown("""
| 常见错误 | 正确做法 | 影响 |
|----------|---------|------|
| 收入成本混淆（净额入账） | 收入成本分开列示，不可轧差 | 低估收入规模 |
| 管理费用和销售费用混用 | 按费用性质分类 | 利润表结构不准确 |
| 税金及附加忘记录入 | 每季末根据增值税计算附加税 | 成本少计、利润膨胀 |
| 社保费漏记 | 每月社保缴费后及时入账 | 费用不完整 |
| 银行手续费忽略 | 检查每笔银行流水中的手续费 | 财务费用不完整 |
| 利润表与申报表不一致 | 交叉校验后再申报 | 申报风险 |
| 亏损月份不做账 | 亏损也要完整记账零申报 | 税务合规风险 |
            """)

    # ==========================================
    #  四、增值税及附加税等税目
    # ==========================================
    else:
        st.subheader("四、增值税及附加税等税目申报说明")

        with st.expander("4.1 税种总览", expanded=True):
            st.markdown("""
**湖北启贤托儿所有限公司涉及的税种（小规模纳税人）：**

| 序号 | 税种 | 税率/征收率 | 申报周期 | 截止日期 | 优惠政策 |
|------|------|-----------|---------|---------|---------|
| 1 | 增值税 | 1%（减按） | 季度 | 季后15日 | 季≤30万免税 |
| 2 | 城建税 | 3.5%（减半） | 季度 | 随增值税 | 六税两费减半 |
| 3 | 教育费附加 | 1.5%（减半） | 季度 | 随增值税 | 六税两费减半 |
| 4 | 地方教育附加 | 1%（减半） | 季度 | 随增值税 | 六税两费减半 |
| 5 | 企业所得税 | 5%（小微） | 季度预缴+年度 | 季后15日/次年5.31 | 小型微利优惠 |
| 6 | 个人所得税 | 3%-45%累进 | 月度 | 次月15日 | 起征点5000元 |
| 7 | 印花税 | 各税目减半 | 按次/按期 | 发生时 | 资金账簿0.0125%/购销0.015%等 |
| 8 | 残保金 | 免征 | 年度 | 年度 | ≤30人免征 |
            """)

        with st.expander("4.2 增值税申报（小规模纳税人）", expanded=False):
            st.markdown("""
**小规模纳税人增值税政策（2026年·湖北）：**

| 场景 | 征收率 | 政策依据 |
|------|--------|---------|
| 季度不含税收入 ≤ 30万元 | **免征** | 财税〔2023〕19号 |
| 季度不含税收入 > 30万元 | **减按1%**（原3%） | 财税〔2023〕19号 |

**关键概念：**
```
含税收入 → 不含税收入：
  不含税收入 = 含税收入 ÷ (1 + 征收率)

举例：
  季度含税收入 50万元
  不含税收入 = 500,000 ÷ 1.03 ≈ 485,436.89 元
  > 30万元 → 应缴增值税 = 485,436.89 × 1% ≈ 4,854.37 元
```

**申报平台**：湖北省电子税务局 → 「增值税及附加税费申报（小规模纳税人适用）」

**申报步骤：**
```
第1步：登录 → 湖北省电子税务局
第2步：选择 → 「增值税及附加税费申报（小规模纳税人）」
第3步：填报 → 含税销售额 → 系统自动换算不含税收入
第4步：免税判断 → ≤30万自动免税，>30万按1%计税
第5步：附加税 → 系统自动计算城建税+教育费附加+地方教育附加
第6步：核对 → 确认无误后申报并缴款
```

**⚠️ 特别注意：**
- 即使享受免税优惠，也必须完成申报（填报后税额为0）
- 增值税发票需通过税控系统开具（全电发票/电子发票）
- 免税收入对应的附加税同步为零
            """)

        with st.expander("4.3 附加税详解（六税两费减半）", expanded=False):
            st.markdown("""
**附加税 = 以实际缴纳的增值税为计税基础：**

| 附加税种 | 名义税率 | 减半后税率 | 计算公式 |
|----------|---------|-----------|---------|
| 城建税（市区） | 7% | **3.5%** | 增值税 × 3.5% |
| 教育费附加 | 3% | **1.5%** | 增值税 × 1.5% |
| 地方教育附加 | 2% | **1%** | 增值税 × 1% |
| **附加税合计** | **12%** | **6%** | 增值税 × 6% |

**「六税两费」减半政策说明：**

| 要素 | 说明 |
|------|------|
| 适用主体 | 小规模纳税人、小型微利企业、个体工商户 |
| 湖北启贤托儿所 | ✅ 既是小规模纳税人，又是小型微利企业 |
| 减半范围 | 城建税、教育费附加、地方教育附加、印花税、房产税、城镇土地使用税、耕地占用税、资源税 |
| 政策依据 | 财税〔2022〕10号 |
| 有效期 | 至2027年12月31日 |
| 申请方式 | 无需申请，湖北省「免申即享」 |

**举例：**
```
增值税应缴 4,854.37 元

减半前：                   减半后：
  城建税   4,854.37×7% = 339.81    城建税   4,854.37×3.5% = 169.90
  教育附加 4,854.37×3% = 145.63    教育附加 4,854.37×1.5% =  72.82
  地方教育 4,854.37×2% =  97.09    地方教育 4,854.37×1%  =  48.54
  ──────────────────────────────    ──────────────────────────────
  合计                582.53 元     合计                   291.26 元

  减免金额：582.53 - 291.26 = 291.27 元 ✅
```
            """)

        with st.expander("4.4 其他税种说明", expanded=False):
            st.markdown("""
**📜 印花税（《中华人民共和国印花税法》· 2022年7月1日施行）：**

| 税目 | 税率 | 计税基础 | 优惠后（减半） | 说明 |
|------|------|---------|--------------|------|
| **资金账簿** | 0.025%（万分之二点五） | 实收资本 + 资本公积 | **0.0125%** | ⚠️ 注册资本到位/增资时缴纳，已缴部分不重复 |
| 购销合同 | 0.03%（万分之三） | 购销合同金额 | **0.015%** | 按季汇总申报 |
| 借款合同 | 0.005%（十万分之五） | 借款金额 | **0.0025%** | 银行贷款/融资 |
| 技术合同 | 0.03%（万分之三） | 技术合同金额 | **0.015%** | 技术开发/转让/咨询 |
| 财产租赁合同 | 0.1%（千分之一） | 租赁金额 | **0.05%** | 房屋/设备租赁 |
| 其他账簿 | 免征 | — | 免征 | 财税〔2018〕50号，已免 |

**⚠️ 资金账簿印花税详解（最重要）：**

| 场景 | 计税基础 | 举例 |
|------|---------|------|
| 公司成立·注册资本首次到位 | 实收资本 + 资本公积全额 | 注册资本100万，实缴到位 → 100万 × 0.0125% = **125元** |
| 后续增资 | 仅对**新增实收资本**部分 | 从100万增资到200万 → 100万 × 0.0125% = **125元** |
| 注册资本未全额到位 | 按实际到位金额 | 注册资本100万，只到位50万 → 50万 × 0.0125% = **62.5元** |
| 已缴过印花税的资本 | **不重复征收** | 无需再次缴纳 |

- 资金账簿印花税在电子税务局「印花税申报」中选择「资金账簿」税目
- 增资变更后需在工商变更登记完成后的纳税期内申报
- 申报方式：按次或按期汇总
- 六税两费减半后，以上各税目均按半额征收
- 小额印花税可通过电子税务局「简并申报」功能汇总缴纳

**残疾人就业保障金（残保金）：**

| 要素 | 说明 |
|------|------|
| 计算公式 | （上年职工人数×1.5% - 实际残疾职工数）× 上年职工平均工资 |
| 免征条件 | 在职职工 ≤ 30人 → 免征 ✅ |
| 政策依据 | 发改价格规〔2019〕2015号 |
| 申报时间 | 年度（通常每年7-9月） |

湖北启贤托儿所在职职工 ≤ 30人，符合免征条件，只需进行「零申报」即可。

**房产税 / 城镇土地使用税：**
- 如公司名下无自有房产/土地，无需缴纳
- 如租赁办公场所，由出租方缴纳，承租方无需申报
            """)

        with st.expander("4.5 优惠政策总览（湖北省2026）", expanded=False):
            st.markdown("""
**湖北启贤托儿所有限公司当前适用的全部税收优惠：**

| 序号 | 税种 | 优惠名称 | 优惠内容 | 政策依据 | 有效期 |
|------|------|---------|---------|---------|--------|
| 1 | 增值税 | 小规模纳税人免税 | 季≤30万免税 | 财税〔2023〕19号 | 至2027.12.31 |
| 2 | 增值税 | 减按1%征收 | 季>30万按1% | 财税〔2023〕19号 | 至2027.12.31 |
| 3 | 城建税 | 六税两费减半 | 7%→3.5% | 财税〔2022〕10号 | 至2027.12.31 |
| 4 | 教育费附加 | 六税两费减半 | 3%→1.5% | 财税〔2022〕10号 | 至2027.12.31 |
| 5 | 地方教育附加 | 六税两费减半 | 2%→1% | 财税〔2022〕10号 | 至2027.12.31 |
| 6 | 企业所得税 | 小型微利优惠 | 实际税负5% | 财税〔2023〕12号 | 至2027.12.31 |
| 7 | 印花税 | 小型微利减半 | 各税目减半 | 财税〔2022〕10号 | 至2027.12.31 |
| 8 | 残保金 | 小微企业免征 | ≤30人免缴 | 发改价格规〔2019〕2015号 | 长期 |

> 🎯 **湖北省特色**：「免申即享」—— 以上所有优惠在申报时系统自动判定和减免，无需额外提交申请或备案材料。
            """)

        with st.expander("4.6 申报日历与缴款流程", expanded=False):
            st.markdown("""
**月度/季度申报时间线：**

```
每月 1~15日：
  └─ 个人所得税（代扣代缴）— 自然人电子税务局扣缴端

每季度结束后 15日内：
  ├─ 增值税及附加税 — 湖北省电子税务局
  ├─ 企业所得税预缴 — 湖北省电子税务局
  └─ 印花税（如有）— 湖北省电子税务局

每年 1月1日~6月30日：
  └─ 工商年报 — 国家企业信用信息公示系统

每年 5月31日前：
  └─ 企业所得税年度汇算清缴 — 湖北省电子税务局

每年 7~9月：
  └─ 残保金申报（零申报）— 湖北省电子税务局
```

**缴款方式：**

| 方式 | 说明 | 推荐 |
|------|------|------|
| 三方协议扣款 | 企业-银行-税务三方签约，自动扣款 | ⭐ 推荐 |
| 银行端查询缴税 | 在电子税务局生成缴款书，到银行柜台缴款 | 备用 |
| 网银直接缴款 | 部分银行支持电子税务局内嵌缴款 | 可选 |

**申报后的检查清单：**
- [ ] 申报状态：申报成功
- [ ] 缴款状态：已缴款（或零申报无需缴款）
- [ ] 回执下载：已留存PDF
- [ ] 申报表打印归档
- [ ] 如涉及退税，跟踪退税进度
            """)

    # ===== 底部通用下载 =====
    st.divider()
    st.subheader("📥 下载完整申报指南")

    # 生成完整指南文本
    full_guide = """================================================================
  湖北启贤托儿所有限公司 · 税务申报操作指南（2026年度）
  基于小规模纳税人 + 小型微利企业 + 湖北省优惠政策
================================================================

一、个人所得税申报（代扣代缴）
================================================================

1. 申报主体：湖北启贤托儿所有限公司（扣缴义务人）
2. 申报周期：月度
3. 申报平台：自然人电子税务局（扣缴端）
4. 截止日期：次月15日

【计算公式（累计预扣法）】
  应税收入 = 税前工资 - 个人社保 - 5000（起征点） - 专项附加扣除
  应纳税额 = 应税收入 × 税率 - 速算扣除数

【7级累进税率】
  级数  应纳税所得额      税率  速算扣除数
  1     ≤3,000            3%    0
  2     3,001~12,000      10%   210
  3     12,001~25,000     20%   1,410
  4     25,001~35,000     25%   2,660
  5     35,001~55,000     30%   4,410
  6     55,001~80,000     35%   7,160
  7     >80,000           45%   15,160

【专项附加扣除（每月标准）】
  子女教育：    2,000元/人
  婴幼儿照护：  2,000元/人
  赡养老人：    最高3,000元/月
  住房贷款利息： 1,000元/月
  住房租金：    1,500元/月（武汉）
  继续教育：    400元/月

【操作步骤】
  1. 人员信息采集（新增/修改员工）
  2. 专项附加扣除下载更新
  3. 综合所得申报 → 正常工资薪金 → 税款计算 → 申报表报送
  4. 税款缴纳（零申报也需完成报送）

【零申报】
  所有员工应税收入为0时仍需完成申报，不可遗漏。


二、企业所得税季报 & 年报
================================================================

【季度预缴（A200000表）】
  申报周期：每季度一次
  截止日期：Q1→4/15, Q2→7/15, Q3→10/15, Q4→次年1/15
  申报平台：湖北省电子税务局

  A200000关键行次：
    第1行   营业收入
    第2行   营业成本
    第3行   利润总额
    第8行   实际利润额
    第10行  应纳所得税额 = 第8行 × 25%
    第11行  减免所得税额（小微优惠）
    第13行  本期应补(退)税额

【小型微利企业优惠】
  条件：年利润≤300万 + 员工≤300 + 资产≤5000万
  实际税率：5%（减按25%×20%）
  方式：系统自动判定，无需申请（免申即享）

【年度汇算清缴】
  截止日期：次年5月31日
  核心逻辑：
    会计利润总额
    + 纳税调增项
    - 纳税调减项
    - 弥补亏损
    = 应纳税所得额 × 5%
    - 已预缴税额
    = 应补(退)税额

【工商年报】
  平台：国家企业信用信息公示系统
  时间：每年1月1日~6月30日
  内容：基本信息、股东出资、资产状况、社保信息等


三、利润表编制说明（小企业会计准则）
================================================================

【利润表标准格式】
  第1行   一、营业收入
  第2行     减：营业成本
  第3行     减：税金及附加
  第5行     减：管理费用（含工资社保等）
  第6行     减：财务费用
  第8行     加：投资收益
  第9行   二、营业利润 = 1-2-3-5-6+8
  第10行    加：营业外收入
  第11行    减：营业外支出
  第12行  三、利润总额 = 9+10-11
  第13行    减：所得税费用 = 12×5%（小型微利）
  第14行  四、净利润 = 12-13

【银行流水 → 利润表映射】
  货款/销售收入/服务费     → 营业收入（第1行）
  采购/进货/材料成本       → 营业成本（第2行）
  税金/城建税/印花税       → 税金及附加（第3行）
  工资/社保/办公/差旅/招待 → 管理费用（第5行）
  利息/银行手续费          → 财务费用（第6行）
  政府补助/罚款收入        → 营业外收入（第10行）
  罚款支出/捐赠            → 营业外支出（第11行）

【关键校验】
  利润表第1行（营业收入）= 申报表A200000 第1行
  利润表第2行（营业成本）= 申报表A200000 第2行
  利润表第12行（利润总额）= 申报表A200000 第3行


四、增值税及附加税等税目申报
================================================================

【税种总览（小规模纳税人）】
  税种            税率        周期    截止      优惠
  增值税          1%（减按）  季度    季后15日  季≤30万免税
  城建税          3.5%（减半）季度    随增值税  六税两费减半
  教育费附加      1.5%（减半）季度    随增值税  六税两费减半
  地方教育附加    1%（减半）  季度    随增值税  六税两费减半
  企业所得税      5%（小微）  季度+年 季后/5.31  小型微利优惠
  个人所得税      3%-45%      月度    次月15日   起征点5000
  印花税          0.03%减半   按次    -         小微减半
  残保金          免征        年度    -         ≤30人免征

【增值税计算】
  不含税收入 = 含税收入 ÷ (1+征收率)
  季≤30万 → 免税
  季>30万 → 应缴 = 不含税收入 × 1%

【附加税 = 增值税 × 优惠后税率】
  城建税      3.5%（原7%减半）
  教育费附加  1.5%（原3%减半）
  地方教育附加 1%（原2%减半）
  合计        6%（原12%减半）

【六税两费减半政策】
  适用：小规模纳税人 + 小型微利企业 ✅
  范围：城建税、教育费附加、地方教育附加、印花税、房产税、
        城镇土地使用税、耕地占用税、资源税
  依据：财税〔2022〕10号
  期限：至2027年12月31日

【申报平台汇总】
  个人所得税：     自然人电子税务局（扣缴端）
  增值税及附加：   湖北省电子税务局
  企业所得税：     湖北省电子税务局
  印花税：         湖北省电子税务局
  残保金：         湖北省电子税务局
  工商年报：       国家企业信用信息公示系统

【湖北省优惠政策总览】
  1. 增值税：小规模减按1%，季≤30万免税
  2. 六税两费：8项减半征收
  3. 企业所得税：小型微利实际5%
  4. 残保金：≤30人免征
  5. 方式：免申即享，系统自动减免

================================================================
  生成时间：""" + datetime.now().strftime("%Y-%m-%d %H:%M") + """
  启贤托育AI税务助手 · 仅供参考，以税务机关最新公告为准
================================================================"""

    dl_g1, dl_g2 = st.columns(2)
    with dl_g1:
        st.download_button(
            label="📥 下载完整申报指南（TXT）",
            data=full_guide,
            file_name=f"税务申报操作指南_湖北启贤托儿所_{datetime.now().strftime('%Y%m%d')}.txt",
            mime="text/plain",
            use_container_width=True,
        )
    with dl_g2:
        pdf_bytes = make_pdf(
            "税务申报操作指南 - 湖北启贤托儿所有限公司",
            full_guide.split("\n"),
            ""
        )
        if pdf_bytes:
            st.download_button(
                label="📥 下载完整申报指南（PDF）",
                data=pdf_bytes,
                file_name=f"税务申报操作指南_湖北启贤托儿所_{datetime.now().strftime('%Y%m%d')}.pdf",
                mime="application/pdf",
                use_container_width=True,
            )

# ===============================================
#  Tab7：残疾人就业保障金申报
# ===============================================

with tab5:
    st.header("♿ 残疾人就业保障金（残保金）申报")
    st.caption("湖北启贤托儿所有限公司 · 小规模纳税人 · 小型微利企业")

    # ── 读取侧边栏参数 ──
    def_prev_employees = st.session_state.get("def_prev_employees", 5)
    def_prev_disabled = st.session_state.get("def_prev_disabled", 0)
    def_prev_avg_salary = st.session_state.get("def_prev_avg_salary", 60000.0)
    def_local_avg_salary = st.session_state.get("def_local_avg_salary", 90000.0)
    def_year = st.session_state.get("def_year", 2026)

    # ===== 计算残保金 =====
    fund_data = calc_disabled_employment_fund(
        prev_year_employees=def_prev_employees,
        prev_year_disabled_employees=def_prev_disabled,
        prev_year_avg_salary=def_prev_avg_salary,
        local_avg_salary=def_local_avg_salary,
        year=def_year,
    )

    # ===== 政策速览区 =====
    col_a, col_b = st.columns(2)
    with col_a:
        st.metric("👥 上年在职职工", f"{fund_data['上年职工人数']} 人")
        st.metric("🧑‍🦽 安排残疾人", f"{fund_data['上年残疾职工人数']} 人")
        st.metric("📐 法定应安排比例", "1.5%")
    with col_b:
        st.metric("💰 职工年均工资", f"{fund_data['上年职工年均工资']:,.0f} 元")
        st.metric("👤 应安排人数", f"{fund_data['应安排人数']:.2f} 人")
        if fund_data["是否小微企业免征"] == "是 ✅":
            st.metric("🎉 应缴残保金", "0 元（免征）", delta="全额减免")
        else:
            st.metric("📋 应缴残保金", f"{fund_data['应缴残保金']:,.2f} 元")

    st.divider()

    # ===== Part 1：小微企业免征判断 =====
    st.subheader("1️⃣ 小微企业免征判定")

    if fund_data["是否小微企业免征"] == "是 ✅":
        st.success(
            f"✅ **免征残疾人就业保障金**\n\n"
            f"| 条件 | 数据 | 判定 |\n"
            f"|------|------|------|\n"
            f"| 在职职工人数 | **{fund_data['上年职工人数']} 人** | ≤ 30人 ✅ |\n"
            f"| 政策依据 | 发改价格规〔2019〕2015号 | — |\n\n"
            f"**结论：您无需缴纳残保金。** 但需要在规定时间内登录电子税务局完成「零申报」。"
        )
    else:
        st.warning(
            f"⚠️ 在职职工 {fund_data['上年职工人数']} 人 > 30人，不符合小微企业免征条件，需按规定计算缴纳。"
        )

    # ===== Part 2：计算明细 =====
    st.subheader("2️⃣ 残保金计算明细")

    col1, col2 = st.columns(2)

    with col1:
        st.markdown("""
**基础公式：**
```
保障金 = (上年职工人数 × 1.5% - 上年残疾职工人数)
       × 上年职工年平均工资
       × 分档征收比例
```

**参数代入：**
| 参数 | 数值 |
|------|------|
| (A) 上年在职职工人数 | {} 人 |
| (B) 法定比例 | 1.5% |
| (C) 应安排人数 = A×B | {:.2f} 人 |
| (D) 实际安排残疾人 | {} 人 |
| (E) 差额人数 = C-D | {:.2f} 人 |
| (F) 职工年均工资 | {:,.0f} 元 |
| (G) 工资基数（封顶后） | {:,.0f} 元 |
| (H) 分档征收比例 | {} |
""".format(
    fund_data["上年职工人数"],
    fund_data["应安排人数"],
    fund_data["上年残疾职工人数"],
    fund_data["差额人数"],
    fund_data["上年职工年均工资"],
    fund_data["工资计算基数"],
    fund_data["分档征收比例"],
))

    with col2:
        if fund_data["应缴残保金"] > 0:
            st.markdown(f"""
**逐步计算：**

1. 应安排人数缺口：
   `{fund_data['上年职工人数']} × 1.5% - {fund_data['上年残疾职工人数']} = {fund_data['差额人数']:.2f} 人`

2. 工资基数（社平2倍封顶）：
   `min({fund_data['上年职工年均工资']:,.0f}, 社平工资×2) = {fund_data['工资计算基数']:,.0f} 元`
   {fund_data['工资封顶说明']}

3. 全额应缴：
   `{fund_data['差额人数']:.2f} × {fund_data['工资计算基数']:,.0f} = {fund_data['应缴残保金（全额）']:,.2f} 元`

4. {fund_data['分档说明']}

5. 实际应缴：
   `{fund_data['应缴残保金（全额）']:,.2f} × {fund_data['分档征收比例']} = **{fund_data['应缴残保金']:,.2f} 元**`

6. 减免金额：
   `{fund_data['应缴残保金（全额）']:,.2f} - {fund_data['应缴残保金']:,.2f} = {fund_data['减免金额']:,.2f} 元` ✅
""")
        else:
            st.info(
                f"**{fund_data['计算说明']}**\n\n"
                f"小微企业免征，应缴金额为 0 元。"
            )

    st.divider()

    # ===== Part 3：优惠政策全览 =====
    st.subheader("3️⃣ 湖北省/武汉市残保金优惠政策（2026）")

    policy_cols = st.columns(3)
    with policy_cols[0]:
        st.markdown("""
**🎯 免征条件（全额）**

| 条件 | 适用 |
|------|------|
| 在职职工 ≤ 30人 | ✅ |
| 安排残疾人比例 ≥ 1.5% | — |

""")
        if def_prev_employees <= 30:
            st.success("✅ 满足免征条件")
        else:
            st.info("不满足免征条件")

    with policy_cols[1]:
        st.markdown("""
**📉 分档减征（>30人）**

| 安排比例 | 征收比例 |
|----------|---------|
| ≥ 1.5% | 免征 |
| 1%~1.5% | 减按 50% |
| < 1% | 按 90% |

""")

    with policy_cols[2]:
        st.markdown("""
**💰 工资基数封顶**

| 项目 | 说明 |
|------|------|
| 封顶线 | 当地社平 × 2 |
| 超封顶 | 按封顶计算 |
| 工资口径 | 应发工资 |

""")

    st.markdown("""
| 序号 | 优惠政策 | 内容 | 政策依据 | 有效期 |
|------|---------|------|---------|--------|
| 1 | 小微企业免征 | 在职职工 ≤ 30人，暂免征收 | 发改价格规〔2019〕2015号 | 长期 |
| 2 | 分档减征 | 达标1%减半、不足1%按90% | 财税〔2019〕98号 | 至2027.12.31 |
| 3 | 工资基数封顶 | 超社平2倍部分不计入 | 财税〔2018〕39号 | 长期 |
| 4 | 湖北省延续 | 上述优惠延续执行 | 湖北省2025助企措施 | 至2027.12.31 |
""")

    # ===== Part 4：申报流程 =====
    st.subheader("4️⃣ 申报操作流程")

    st.markdown("""
**申报平台**：湖北省电子税务局 → 「非税收入通用申报」或「残疾人就业保障金申报」

**申报时间**：每年 **7月1日 ~ 9月30日**（以当地税务机关公告为准）

**操作步骤（零申报场景）：**

```
第1步：登录 → 湖北省电子税务局 → 企业登录
第2步：进入 → 「我要办税」 → 「税费申报及缴纳」
第3步：查找 → 「非税收入通用申报」或搜索「残保金」
第4步：填写 → 上年在职职工人数、安排残疾人数、年均工资
第5步：系统自动判断 → ≤30人自动免征（应缴=0）
第6步：申报 → 确认无误后点击「申报」
第7步：截止 → 即使为零也需要完成报送
第8步：留存 → 下载申报回执PDF归档
```

**操作步骤（需缴款场景，>30人）：**

```
第1-4步：同上
第5步：系统自动计算应缴金额
第6步：核对计算结果是否正确
第7步：申报 → 确认无误后点击「申报」
第8步：缴款 → 通过三方协议扣款或银行端缴款
第9步：留存 → 下载完税凭证PDF → 财务记账
```

**⚠️ 注意事项：**
- **上年数据** = 申报年度的上一年（如2026年申报，填写2025年数据）
- **在职职工人数** = 全年各月平均人数（含季节性用工需折算）
- **残疾职工人数** = 须经残联审核确认的残疾人就业人数
- 如公司无残疾职工，填写 0 人
- 逾期不申报将被认定为未安置残疾人，按全额计算并可能加收滞纳金
""")

    # ===== Part 5：常见问题 =====
    st.subheader("5️⃣ 常见问题")

    with st.expander("Q1：湖北启贤托儿所需要缴纳残保金吗？"):
        st.markdown(f"""
**答：不需要（免征）。**

湖北启贤托儿所有限公司在职职工 **{def_prev_employees} 人 ≤ 30人**，符合小微企业免征条件（发改价格规〔2019〕2015号）。

但必须每年在规定时间内（7~9月）登录电子税务局完成零申报手续，不可遗漏。
        """)

    with st.expander("Q2：零申报也必须做吗？不做会怎样？"):
        st.markdown("""
**答：必须做。**

即使应缴金额为0，也必须按时完成申报。逾期未申报将被视为：
- 未按规定安排残疾人就业
- 可能被按全额计算并追缴残保金
- 纳入信用记录
- 影响企业纳税信用等级
        """)

    with st.expander("Q3：员工人数怎么算？包含老板吗？"):
        st.markdown("""
**答：包含所有在职职工。**

- 「在职职工」= 与单位签订劳动合同、由单位支付工资的所有人员
- 包括：正式员工、合同工、季节性用工、临时工（不含劳务派遣）
- **老板（法定代表人）如从公司领取工资，也计入在内**
- 计算方式：全年各月平均 = 各月月末人数之和 ÷ 12
- 季节性用工需折算为年平均用工人数
        """)

    with st.expander("Q4：可以聘用残疾人来抵扣残保金吗？"):
        st.markdown("""
**答：可以。**

安排残疾人就业是减免残保金的最有效方式：

| 安排人数 | 效果 |
|----------|------|
| 安排1名重度残疾人 | 按2名计算 |
| 达到1.5%比例 | 全额免征 |

但需注意：
- 残疾职工须持有《残疾人证》
- 需签订1年以上劳动合同
- 需足额缴纳社保
- 需到当地残联办理「按比例就业审核」确认
        """)

    with st.expander("Q5：去年刚成立的公司要不要申报？"):
        st.markdown("""
**答：要。**

- 成立不足1年的，按实际月份计算平均人数
- 仍适用≤30人免征政策
- 第一次申报时需先在电子税务局做税费种认定（如未自动带出）

**计算公式**（成立不足1年）：
```
月平均人数 = 各月人数之和 ÷ 实际经营月数
```
        """)

    with st.expander("Q6：残保金和社保里的工伤保险是一回事吗？"):
        st.markdown("""
**答：不是。**

| 项目 | 性质 | 征收部门 | 用途 |
|------|------|---------|------|
| 工伤保险 | 社保五险之一 | 社保局 | 因工受伤赔偿 |
| 残保金 | 政府性基金 | 税务代征→残联使用 | 残疾人就业培训、补贴 |

两者是完全不同的款项，不可混淆。即使缴纳残保金，也不能替代工伤保险。
        """)

    # ===== 底部下载 =====
    st.divider()
    st.subheader("📥 下载残保金测算底稿")

    # 生成通知文本
    if fund_data["应缴残保金"] == 0:
        status_text = "零申报（免征）"
        amount_text = "0 元"
        notice_text = "贵公司符合小微企业免征条件（在职职工≤30人），应缴金额为0元。请在规定期限内完成零申报即可。"
    else:
        status_text = f"需缴纳 {fund_data['应缴残保金']:,.2f} 元"
        amount_text = f"{fund_data['应缴残保金']:,.2f} 元"
        notice_text = f"请于 {fund_data['申报年度']} 年申报期内完成申报及缴款。"

    fund_report = f"""================================================================
       残疾人就业保障金测算报告
  填报单位：湖北启贤托儿所有限公司
  申报年度：{fund_data['申报年度']}年
================================================================

一、基本信息
  上年（{fund_data['申报年度']-1}年）在职职工人数：{fund_data['上年职工人数']} 人
  上年安排残疾人就业人数：        {fund_data['上年残疾职工人数']} 人
  上年职工年平均工资：            {fund_data['上年职工年均工资']:,.2f} 元
  法定安排比例：                  {fund_data['法定安排比例']}
  应安排残疾人数：                {fund_data['应安排人数']:.4f} 人

二、优惠政策适用
  {fund_data['计算说明']}
  政策依据：{fund_data['政策依据']}

三、计算结果
  应缴残保金（全额）：{fund_data['应缴残保金（全额）']:,.2f} 元
  分档征收比例：      {fund_data['分档征收比例']}
  减免金额：          {fund_data['减免金额']:,.2f} 元
  ─────────────────────────────
  实际应缴残保金：    {fund_data['应缴残保金']:,.2f} 元

四、申报结论
  状态：{status_text}
  {notice_text}
  申报截止：{fund_data['申报截止']}

五、重要提示
  1. 即使免征也必须按时在电子税务局完成零申报；
  2. 上年数据需与上一年度个税申报人数一致，税务系统会自动比对；
  3. 如安排有残疾人，需在申报前完成残联审核认定；
  4. 请保存本测算报告和申报回执至少5年。

================================================================
  生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M')}
  启贤托育AI税务助手 · 仅供参考，以税务机关最新公告为准
================================================================"""

    dl_f1, dl_f2 = st.columns(2)
    with dl_f1:
        st.download_button(
            label=f"📥 下载残保金测算报告（TXT）",
            data=fund_report,
            file_name=f"残保金测算报告_{def_year}年_湖北启贤托儿所.txt",
            mime="text/plain",
            use_container_width=True,
        )
    with dl_f2:
        pdf_bytes = make_pdf(
            f"残疾人就业保障金测算报告 - {def_year}年",
            fund_report.split("\n"),
            ""
        )
        if pdf_bytes:
            st.download_button(
                label=f"📥 下载残保金测算报告（PDF）",
                data=pdf_bytes,
                file_name=f"残保金测算报告_{def_year}年_湖北启贤托儿所.pdf",
                mime="application/pdf",
                use_container_width=True,
            )

# ===============================================
#  Tab8：年报数据导入
# ===============================================

with tab1:
    st.header("🗂️ 年报数据导入")
    st.caption("支持 Excel / PDF 两种格式。导入后年报数据自动拆分为 4 个季度申报底稿。税务年报与内部底稿不一致时，重新导入即可纠偏。")
    st.success("✅ v1.9.0 — 合并个税计算Tab（单人+批量双模式）（2026-06-03 build）")

    # ── 检查是否有历史导入 ──
    snapshot_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), "年报导入快照.json")
    prev_snapshot = None
    if os.path.exists(snapshot_file):
        try:
            with open(snapshot_file, "r", encoding="utf-8") as f:
                prev_snapshot = json.load(f)
        except Exception:
            prev_snapshot = None

    if prev_snapshot:
        prev_summary = prev_snapshot.get("summary", {})
        st.info(f"""
        📌 **已有导入记录**（{prev_snapshot.get('imported_at', '未知时间')}）
        — 营业收入 {prev_summary.get('annual_revenue', 0):,.0f} 元 |
        利润 {prev_summary.get('annual_profit', 0):,.0f} 元 |
        {prev_snapshot.get('employee_count', 0)} 名员工
        — 下方上传新数据将进入「纠偏模式」，可对比差异后覆盖更新。
        """)

    # ── Step 1：下载模板 ──
    st.subheader("📥 第一步：下载导入模板")
    st.markdown("""
    模板包含 3 个 Sheet：
    - **年报汇总** — 全年收入/成本/利润/注册资本（部分实缴）/残保金参数
    - **员工信息** — 每位员工的月均工资和专项附加扣除
    - **季度分摊明细** — 可选，如需按季度不均分则填写
    """)

    template_bytes = gen_annual_report_template_bytes()
    pdf_template_bytes = gen_annual_report_template_pdf_bytes()

    dl_t1, dl_t2 = st.columns(2)
    with dl_t1:
        st.download_button(
            label="📥 下载模板（Excel · 可填写上传）",
            data=template_bytes,
            file_name="年报数据导入模板.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    with dl_t2:
        if pdf_template_bytes:
            st.download_button(
                label="📥 下载模板（PDF · 打印存档）",
                data=pdf_template_bytes,
                file_name="年报数据导入模板.pdf",
                mime="application/pdf",
                use_container_width=True,
            )
        else:
            st.download_button(
                label="📥 PDF 不可用（需 fpdf2）",
                data=b"",
                file_name="",
                mime="text/plain",
                use_container_width=True,
                disabled=True,
            )

    st.divider()

    # ── Step 2：上传 Excel 或 PDF ──
    st.subheader("📤 第二步：上传年报文件（Excel 或 PDF）")

    up_col1, up_col2 = st.columns(2)
    with up_col1:
        uploaded_excel = st.file_uploader(
            "上传 Excel 年报",
            type=["xlsx", "xls"],
            key="annual_report_excel_uploader",
        )
    with up_col2:
        uploaded_pdf = st.file_uploader(
            "上传 PDF 年报（税务返回版）",
            type=["pdf"],
            key="annual_report_pdf_uploader",
        )

    uploaded = uploaded_excel or uploaded_pdf
    is_pdf = uploaded_pdf is not None

    if uploaded is not None:
        fname = uploaded.name
        fsize = uploaded.size
        cache_key = f"parsed_annual_{fname}_{fsize}"
        if cache_key not in st.session_state or st.button("🔄 重新解析", key="reparse_annual"):
            with st.spinner(f"正在{'AI ' if is_pdf else ''}解析年报数据（{fname}）..."):
                file_bytes = uploaded.getvalue()
                if is_pdf:
                    st.session_state[cache_key] = parse_annual_report_pdf(file_bytes)
                else:
                    st.session_state[cache_key] = parse_annual_report_excel(file_bytes)
            st.rerun()

        parsed = st.session_state[cache_key]
        summary = parsed["summary"]
        employees = parsed["employees"]
        quarterly = parsed["quarterly"]
        warnings = parsed["warnings"]

        if is_pdf:
            st.caption("🤖 已通过 AI 从 PDF 中提取数据，请仔细核对各字段是否正确")

        # 显示警告
        if warnings:
            with st.expander(f"🔍 数据校验（{len(warnings)} 条提示）", expanded=True):
                for w in warnings:
                    st.warning(w)

        # ── 纠偏对比（有历史数据时） ──
        if prev_snapshot:
            st.subheader("🔍 纠偏对比：新数据 vs 历史导入")
            prev_s = prev_snapshot.get("summary", {})
            diff_fields = [
                ("annual_revenue", "全年营业收入（含税）", "元"),
                ("annual_cost", "全年营业成本", "元"),
                ("annual_profit", "全年利润总额", "元"),
                ("annual_vat_revenue", "增值税计税收入（不含税）", "元"),
                ("avg_employees", "平均从业人数", "人"),
                ("avg_assets", "平均资产总额", "万元"),
                ("reg_capital", "注册资本实缴额", "元"),
                ("capital_increase", "本年增资额", "元"),
                ("total_salary", "全年工资总额", "元"),
                ("prev_employees", "上年职工人数", "人"),
            ]
            diff_rows = []
            has_diff = False
            for key, label, unit in diff_fields:
                old_val = prev_s.get(key, 0) or 0
                new_val = summary.get(key, 0) or 0
                try:
                    old_val = float(old_val)
                    new_val = float(new_val)
                except (ValueError, TypeError):
                    old_val = 0
                    new_val = 0
                delta = new_val - old_val
                if abs(delta) > 0.01:
                    has_diff = True
                    diff_rows.append({
                        "项目": label,
                        "历史值": f"{old_val:,.2f} {unit}",
                        "新值": f"{new_val:,.2f} {unit}",
                        "差异": f"{delta:+,.2f} {unit}",
                    })
            if has_diff:
                st.dataframe(pd.DataFrame(diff_rows), use_container_width=True, hide_index=True)
                st.warning("⚠️ 税务年报与内部底稿存在差异，确认后将用新数据覆盖旧数据（含季度申报数据）")
            else:
                st.success("✅ 新数据与历史数据一致，无需纠偏")

        # ── 解析结果预览 ──
        st.subheader("📋 第三步：确认解析结果")

        prev_col1, prev_col2, prev_col3 = st.columns(3)

        with prev_col1:
            st.markdown("**📊 经营数据**")
            st.metric("全年营业收入", f"{summary.get('annual_revenue', 0):,.0f} 元")
            st.metric("全年营业成本", f"{summary.get('annual_cost', 0):,.0f} 元")
            st.metric("全年利润总额", f"{summary.get('annual_profit', 0):,.0f} 元")
            st.metric("增值税计税收入", f"{summary.get('annual_vat_revenue', 0):,.0f} 元")

        with prev_col2:
            st.markdown("**🏢 企业信息**")
            st.metric("平均从业人数", f"{int(summary.get('avg_employees', 0))} 人")
            st.metric("平均资产总额", f"{summary.get('avg_assets', 0):,.1f} 万元")
            st.metric("注册资本实缴", f"{summary.get('reg_capital', 0):,.0f} 元")
            st.metric("本年增资额", f"{summary.get('capital_increase', 0):,.0f} 元")
            q_split = summary.get("split_method", "平均")
            st.caption(f"季度分摊方式：**{q_split}**")

        with prev_col3:
            st.markdown("**👥 薪酬 & 残保金**")
            st.metric("全年工资总额", f"{summary.get('total_salary', 0):,.0f} 元")
            st.metric("社保公司承担", f"{summary.get('total_si_company', 0):,.0f} 元")
            st.metric("上年职工人数", f"{int(summary.get('prev_employees', 0))} 人")
            st.metric("上年安排残疾人", f"{int(summary.get('prev_disabled', 0))} 人")

        # 员工预览
        if employees:
            st.markdown(f"**👥 员工信息（{len(employees)} 人）**")
            emp_df = pd.DataFrame(employees)
            emp_df_display = emp_df.rename(columns={
                "name": "姓名", "gross_salary": "税前月工资",
                "si_base": "社保基数", "si_personal_actual": "个人社保",
                "special_deductions": "专项扣除", "child_education": "子女教育",
                "infant_care": "婴幼儿照护", "elderly_care": "赡养老人",
            })
            st.dataframe(emp_df_display, use_container_width=True, hide_index=True)

        # 季度明细预览
        if quarterly:
            st.markdown("**📅 季度分摊明细**")
            q_rows = []
            q_total_rev = q_total_cost = q_total_profit = q_total_vat = 0
            for q in ["Q1", "Q2", "Q3", "Q4"]:
                qd = quarterly.get(q, {})
                r = qd.get("revenue", 0)
                c = qd.get("cost", 0)
                p = qd.get("period_profit", 0)
                v = qd.get("vat_revenue", 0)
                q_total_rev += r
                q_total_cost += c
                q_total_profit += p
                q_total_vat += v
                q_rows.append({
                    "季度": q, "营业收入": r, "营业成本": c, "利润": p, "增值税收入": v,
                })
            # 合计行
            q_rows.append({
                "季度": "合计", "营业收入": q_total_rev, "营业成本": q_total_cost,
                "利润": q_total_profit, "增值税收入": q_total_vat,
            })
            st.dataframe(pd.DataFrame(q_rows), use_container_width=True, hide_index=True)

            # 校验：四季度合计 = 年报总数
            annual_rev = summary.get("annual_revenue", 0)
            annual_cost = summary.get("annual_cost", 0)
            annual_profit = summary.get("annual_profit", 0)
            annual_vat = summary.get("annual_vat_revenue", 0)

            rev_ok = abs(q_total_rev - annual_rev) < 1
            cost_ok = abs(q_total_cost - annual_cost) < 1
            profit_ok = abs(q_total_profit - annual_profit) < 1
            vat_ok = abs(q_total_vat - annual_vat) < 1

            if not (rev_ok and cost_ok and profit_ok and vat_ok):
                issues = []
                if not rev_ok:
                    issues.append(f"营业收入：四季度合计 {q_total_rev:,.0f} ≠ 年报 {annual_rev:,.0f}")
                if not cost_ok:
                    issues.append(f"营业成本：四季度合计 {q_total_cost:,.0f} ≠ 年报 {annual_cost:,.0f}")
                if not profit_ok:
                    issues.append(f"利润总额：四季度合计 {q_total_profit:,.0f} ≠ 年报 {annual_profit:,.0f}")
                if not vat_ok:
                    issues.append(f"增值税收入：四季度合计 {q_total_vat:,.0f} ≠ 年报 {annual_vat:,.0f}")
                st.warning("⚠️ 年报汇总 ≠ 四季度合计，请核对：\n" + "\n".join(f"  · {i}" for i in issues))
            else:
                st.success("✅ 年报汇总 = 四季度合计，数据一致")
        else:
            # 显示自动均分预览
            rev = summary.get("annual_revenue", 0)
            cost = summary.get("annual_cost", 0)
            profit = summary.get("annual_profit", 0)
            vat_rev = summary.get("annual_vat_revenue", 0)
            st.caption("💡 未填写季度明细，将按 4 季度平均分摊：")
            q_avg_rows = []
            for q in ["Q1", "Q2", "Q3", "Q4"]:
                q_avg_rows.append({
                    "季度": q, "营业收入": round(rev / 4, 2),
                    "营业成本": round(cost / 4, 2), "利润": round(profit / 4, 2),
                    "增值税收入": round(vat_rev / 4, 2),
                })
            q_avg_rows.append({
                "季度": "合计", "营业收入": rev, "营业成本": cost,
                "利润": profit, "增值税收入": vat_rev,
            })
            st.dataframe(pd.DataFrame(q_avg_rows), use_container_width=True, hide_index=True)

        # ── 工资数据校验（三重） ──
        if employees:
            with st.expander("🔎 工资数据校验（银行流水 vs 个税申报 vs 年报）", expanded=False):
                st.caption("上传银行流水和/或个税申报记录，与导入的员工工资交叉比对")

                vac1, vac2 = st.columns(2)
                with vac1:
                    val_bank_file = st.file_uploader(
                        "上传银行流水（校验工资支出）",
                        type=["csv", "xlsx", "xls"],
                        key="annual_val_bank",
                    )
                with vac2:
                    val_tax_file = st.file_uploader(
                        "上传个税申报记录（校验累计收入）",
                        type=["csv", "xlsx", "xls"],
                        key="annual_val_tax",
                    )

                val_annual_salary = st.number_input(
                    "年报「全年工资总额」（元，选填）",
                    min_value=0.0, value=float(summary.get("total_salary", 0) or 0.0), step=1000.0,
                    key="annual_val_salary_input",
                )

                if st.button("🔍 开始校验", key="run_annual_salary_val", use_container_width=True):
                    bank_df = None
                    tax_df = None

                    if val_bank_file:
                        try:
                            if val_bank_file.name.endswith(".csv"):
                                try:
                                    bank_df = pd.read_csv(val_bank_file, encoding="utf-8-sig")
                                except Exception:
                                    val_bank_file.seek(0)
                                    bank_df = pd.read_csv(val_bank_file, encoding="gbk")
                            else:
                                bank_df = pd.read_excel(val_bank_file)
                        except Exception as e:
                            st.error(f"银行流水读取失败：{e}")

                    if val_tax_file:
                        try:
                            if val_tax_file.name.endswith(".csv"):
                                try:
                                    tax_df = pd.read_csv(val_tax_file, encoding="utf-8-sig")
                                except Exception:
                                    val_tax_file.seek(0)
                                    tax_df = pd.read_csv(val_tax_file, encoding="gbk")
                            else:
                                tax_df = pd.read_excel(val_tax_file)
                        except Exception as e:
                            st.error(f"个税申报记录读取失败：{e}")

                    with st.spinner("正在校验..."):
                        val_result = validate_salary_data(
                            employees=employees,
                            bank_df=bank_df,
                            tax_filing_df=tax_df,
                            annual_total_salary=val_annual_salary if val_annual_salary > 0 else 0.0,
                        )

                    st.divider()

                    # 校验1：银行流水
                    if bank_df is not None:
                        st.markdown("**🏦 校验一：银行流水 vs 系统工资**")
                        bm = val_result.get("bank_match")
                        if isinstance(bm, dict) and bm:
                            c1, c2, c3 = st.columns(3)
                            c1.metric("银行流水工资支出", f"{bm['bank_salary_total']:,.0f} 元")
                            c2.metric("系统年工资合计", f"{bm['sys_annual_total']:,.0f} 元")
                            c3.metric("差异", f"{bm['diff']:+,.0f} 元", delta=f"{bm['diff_pct']:+.1f}%")
                            if bm["match"]:
                                st.success("✅ 银行流水与系统工资一致")
                            else:
                                st.error(f"⚠️ 差异较大！建议核查（识别到 {bm['txn_count']} 条工资类交易）")
                        else:
                            st.warning("未在银行流水中识别到工资类支出（摘要需含「工资」「奖金」「绩效」等关键词）")
                        st.divider()

                    # 校验2：个税申报
                    if tax_df is not None:
                        st.markdown("**📋 校验二：个税申报记录 vs 系统工资**")
                        tm = val_result.get("tax_match", [])
                        if tm:
                            tm_rows = []
                            for r in tm:
                                tm_rows.append({
                                    "姓名": r["name"],
                                    "个税申报累计收入": f"{r['tax_filing_income']:,.0f}",
                                    "系统年工资": f"{r['sys_annual']:,.0f}",
                                    "差异": f"{r['diff']:+,.0f}",
                                    "状态": "✅" if r["match"] else "⚠️",
                                })
                            st.dataframe(pd.DataFrame(tm_rows), use_container_width=True, hide_index=True)
                        else:
                            st.warning("个税申报记录中未找到与系统员工匹配的姓名")
                        st.divider()

                    # 校验3：年报工资总额
                    if val_annual_salary > 0:
                        st.markdown("**📊 校验三：年报工资总额 vs 系统年工资合计**")
                        am = val_result.get("annual_match", {})
                        if am:
                            c1, c2, c3 = st.columns(3)
                            c1.metric("年报工资总额", f"{am['annual_total_salary']:,.0f} 元")
                            c2.metric("系统年工资合计", f"{am['sys_annual_total']:,.0f} 元")
                            c3.metric("差异", f"{am['diff']:+,.0f} 元", delta=f"{am['diff_pct']:+.1f}%")
                            if am["match"]:
                                st.success("✅ 年报工资总额与系统工资一致")
                            else:
                                st.error("⚠️ 年报工资总额与系统年工资合计差异较大！")
                        st.divider()

                    # 汇总
                    warnings = val_result.get("warnings", [])
                    if warnings:
                        with st.expander(f"📋 校验说明（{len(warnings)} 条）", expanded=True):
                            for w in warnings:
                                if "✅" in w:
                                    st.success(w)
                                elif "⚠️" in w or "差" in w:
                                    st.warning(w)
                                else:
                                    st.info(w)

        st.divider()

        # ── Step 4：确认导入 ──
        if prev_snapshot:
            st.subheader("🚀 第四步：确认导入（纠偏模式）")
            st.markdown("""
            点击下方按钮将用新数据**覆盖**以下内容：
            - ✅ 员工信息 → 更新「工资计算」草稿
            - ✅ 4 个季度数据 → **覆盖**「季度申报」存档
            - ✅ 印花税参数 → 更新注册资本/增资
            - ✅ 残保金参数 → 更新上年职工人数/工资等
            """)
            import_label = "✅ 确认导入（覆盖旧数据）"
        else:
            st.subheader("🚀 第四步：确认导入")
            st.markdown("""
            点击下方按钮将一次性完成以下操作：
            - ✅ 将员工信息保存为「工资计算」草稿
            - ✅ 将 4 个季度数据写入「季度申报」存档
            - ✅ 设置侧边栏的印花税（注册资本/增资）和残保金参数
            """)
            import_label = "✅ 确认导入 2025 年年报数据"

        import_warning = st.checkbox(
            "⚠️ 我确认以上数据正确，导入将覆盖现有草稿和季度申报数据",
            key="confirm_import",
        )

        if st.button(import_label, type="primary", use_container_width=True,
                     disabled=not import_warning):
            import_count = 0

            # 1. 保存员工草稿
            if employees:
                EMP_DATA_FILE_ANNUAL = os.path.join(os.path.dirname(os.path.abspath(__file__)), "员工数据_草稿.json")
                with open(EMP_DATA_FILE_ANNUAL, "w", encoding="utf-8") as f:
                    json.dump(employees, f, ensure_ascii=False, indent=2)
                st.session_state["employees_saved"] = employees
                import_count += 1
                st.success(f"✅ 已导入 {len(employees)} 名员工数据")

            # 2. 写入季度申报数据（2025年）
            rev = summary.get("annual_revenue", 0)
            cost = summary.get("annual_cost", 0)
            profit = summary.get("annual_profit", 0)
            vat_rev = summary.get("annual_vat_revenue", 0)
            avg_emp = int(summary.get("avg_employees", 0))
            avg_assets = summary.get("avg_assets", 0.0)

            if quarterly:
                q_data_map = quarterly
            else:
                q_data_map = {}
                for q in ["Q1", "Q2", "Q3", "Q4"]:
                    q_data_map[q] = {
                        "revenue": rev / 4,
                        "cost": cost / 4,
                        "period_profit": profit / 4,
                        "vat_revenue": vat_rev / 4,
                        "avg_employees": avg_emp,
                        "avg_assets": avg_assets,
                    }

            quarter_imported = 0
            target_year = 2025  # 年报数据所属年度
            for q_name, q_data in q_data_map.items():
                q_num = int(q_name.replace("Q", ""))
                save_quarter_data(target_year, q_num, {
                    "revenue": q_data.get("revenue", 0),
                    "cost": q_data.get("cost", 0),
                    "period_profit": q_data.get("period_profit", 0),
                    "vat_revenue": q_data.get("vat_revenue", 0),
                    "num_employees": int(q_data.get("avg_employees", avg_emp)),
                    "total_assets": q_data.get("avg_assets", avg_assets),
                    "saved_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "_source": "年报导入",
                })
                quarter_imported += 1
            import_count += 1
            st.success(f"✅ 已写入 2025 年 {quarter_imported} 个季度申报数据")

            # 3. 设置侧边栏参数
            st.session_state["stamp_reg_capital"] = float(summary.get("reg_capital", 0))
            st.session_state["stamp_capital_increase"] = float(summary.get("capital_increase", 0))
            st.session_state["def_prev_employees"] = int(summary.get("prev_employees", 0))
            st.session_state["def_prev_disabled"] = int(summary.get("prev_disabled", 0))
            st.session_state["def_prev_avg_salary"] = float(summary.get("prev_avg_salary", 60000))
            st.session_state["def_local_avg_salary"] = float(summary.get("local_avg_salary", 90000))

            # 4. 保存年报快照
            snapshot = {
                "imported_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "year": 2025,
                "source": "PDF" if is_pdf else "Excel",
                "summary": {k: (float(v) if isinstance(v, (int, float)) else v) for k, v in summary.items()},
                "employee_count": len(employees),
                "quarter_count": quarter_imported,
            }
            with open(snapshot_file, "w", encoding="utf-8") as f:
                json.dump(snapshot, f, ensure_ascii=False, indent=2)

            if prev_snapshot:
                prev_s = prev_snapshot.get("summary", {})
                old_rev = prev_s.get("annual_revenue", 0) or 0
                new_rev = summary.get("annual_revenue", 0) or 0
                delta_rev = new_rev - old_rev
                st.balloons()
                st.success(f"""
                🎉 **年报纠偏完成！**

                | 项目 | 旧值 | 新值 | 变化 |
                |------|------|------|------|
                | 营业收入 | {old_rev:,.0f} 元 | {new_rev:,.0f} 元 | {delta_rev:+,.0f} 元 |
                | 员工数据 | {prev_snapshot.get('employee_count', 0)} 人 | {len(employees)} 人 | — |
                | 季度申报 | 已覆盖 | 2025 Q1-Q4 | — |
                | 印花税/残保金 | 已更新 | — | — |

                👉 切换到「📊 季度申报」选择 2026 年即可开始本年申报。
                """)
            else:
                st.balloons()
                st.success(f"""
                🎉 **2025 年年报数据导入完成！**

                | 项目 | 状态 |
                |------|------|
                | 员工数据 | {len(employees)} 人已保存 |
                | 季度申报 | 2025 年 Q1-Q4 已写入 |
                | 印花税参数 | 注册资本 {summary.get('reg_capital', 0):,.0f} 元 |
                | 残保金参数 | 上年 {int(summary.get('prev_employees', 0))} 人 |

                👉 现在切换到「📊 季度申报」并选择 2026 年即可开始本年申报。
                """)

# ===============================================
#  全局页脚
# ===============================================
st.divider()
footer_col1, footer_col2, footer_col3 = st.columns([1, 2, 1])
with footer_col2:
    st.caption(
        f"© {datetime.now().year} 启贤托育AI税务助手 v1.6 · "
        "湖北启贤托儿所有限公司 · "
        "基于小规模纳税人+小型微利企业场景 · "
        "仅供参考，以税务机关最新公告为准"
    )
